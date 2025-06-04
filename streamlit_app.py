import streamlit as st
import pandas as pd
import numpy as np
import os
import tempfile
import base64
from PIL import Image
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import io
import time
import logging
from datetime import datetime, timedelta
import fitz  # PyMuPDF
import re # 정규식 추가
import boto3
import json
from botocore.exceptions import ClientError
import hashlib
from pdf2image import convert_from_path  # PDF를 이미지로 변환하기 위한 라이브러리 추가
from openpyxl import Workbook # Workbook 임포트 확인
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows # dataframe_to_rows 임포트 추가
from openpyxl import load_workbook
from functools import lru_cache
import concurrent.futures
from typing import List, Dict

# 로컬 모듈 임포트
import pdf3_module
import data_analyzer
from data_analyzer import get_unique_departments, filter_by_department, load_excel_data

# 앱 설정
st.set_page_config(
    page_title="상계백병원 인수증 & 엑셀 데이터 비교 시스템",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)



# 로깅 설정
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# botocore 로깅 레벨 설정 (DEBUG -> WARNING)
logging.getLogger('botocore').setLevel(logging.ERROR)
logging.getLogger('boto3').setLevel(logging.ERROR)
logging.getLogger('s3transfer').setLevel(logging.ERROR)
logging.getLogger('urllib3').setLevel(logging.ERROR)



# S3 설정을 secrets에서 가져오기
AWS_CONFIG = {
    "aws_access_key_id": st.secrets["aws"]["AWS_ACCESS_KEY_ID"],
    "aws_secret_access_key": st.secrets["aws"]["AWS_SECRET_ACCESS_KEY"],
    "region_name": st.secrets["aws"]["AWS_REGION"]
}
S3_BUCKET = st.secrets["aws"]["S3_BUCKET"]

# S3 클라이언트 설정
def get_s3_client():
    try:
        return boto3.client('s3', **AWS_CONFIG)
    except Exception as e:
        logger.error(f"S3 클라이언트 생성 실패: {e}")
        return None

# S3 디렉토리 구조
S3_DIRS = {
    "EXCEL": "excel/",
    "PDF": "pdf/",
    "EXTRACTED": "extracted/",
    "OCR_RESULTS": "ocr_results/",
    "METADATA": "metadata/",
    "DB": "db/",  # DB 디렉토리 추가
    "RESULTS": "results/",  # 분석 결과 저장 디렉토리 추가
    "PREVIEW_IMAGES": "preview_images/"  # 미리보기 이미지 디렉토리 추가
}

def get_s3_file_modified_time(s3_handler, key):
    """S3 파일의 마지막 수정 시각을 반환 (datetime)"""
    try:
        response = s3_handler.s3_client.head_object(Bucket=s3_handler.bucket, Key=key)
        return response['LastModified']
    except Exception as e:
        return None
    


class S3Handler:
    def __init__(self):
        self.s3_client = get_s3_client()
        if self.s3_client is None:
            raise Exception("S3 클라이언트 초기화 실패")
        self.bucket = S3_BUCKET
        self.dirs = S3_DIRS
        self.image_cache = ImageCache()
    
    def generate_file_key(self, date_str, filename, dir_type):
        """파일 키 생성 (경로)"""
        return f"{self.dirs[dir_type]}{date_str}/{filename}"

    def upload_file(self, file_obj, date_str, original_filename, dir_type):
        """파일 업로드"""
        try:
            file_key = self.generate_file_key(date_str, original_filename, dir_type)
            self.s3_client.upload_fileobj(file_obj, self.bucket, file_key)
            return {"status": "success", "key": file_key}
        except Exception as e:
            logger.error(f"S3 업로드 실패 ({original_filename}): {e}")
            return {"status": "error", "message": str(e)}

    def download_file(self, file_key):
        """파일 다운로드"""
        try:
            response = self.s3_client.get_object(Bucket=self.bucket, Key=file_key)
            return {"status": "success", "data": response['Body'].read()}
        except Exception as e:
            logger.error(f"S3 다운로드 실패 ({file_key}): {e}")
            return {"status": "error", "message": str(e)}

    def get_s3_file_modified_time(s3_handler, key):
        """S3 파일의 마지막 수정 시각을 반환 (datetime)"""
        try:
            response = s3_handler.s3_client.head_object(Bucket=s3_handler.bucket, Key=key)
            return response['LastModified']
        except Exception as e:
            return None

    def save_metadata(self, date_str, metadata):
        """메타데이터 저장"""
        try:
            metadata_key = f"{self.dirs['METADATA']}{date_str}/metadata.json"
            self.s3_client.put_object(
                Bucket=self.bucket,
                Key=metadata_key,
                Body=json.dumps(metadata, ensure_ascii=False)
            )
            return {"status": "success", "key": metadata_key}
        except Exception as e:
            logger.error(f"메타데이터 저장 실패 ({date_str}): {e}")
            return {"status": "error", "message": str(e)}

    def load_metadata(self, date_str):
        """메타데이터 로드"""
        try:
            metadata_key = f"{self.dirs['METADATA']}{date_str}/metadata.json"
            response = self.s3_client.get_object(Bucket=self.bucket, Key=metadata_key)
            return {"status": "success", "data": json.loads(response['Body'].read())}
        except ClientError as e:
            if e.response['Error']['Code'] == 'NoSuchKey':
                return {"status": "not_found"}
            logger.error(f"메타데이터 로드 실패 ({date_str}): {e}")
            return {"status": "error", "message": str(e)}

    def list_processed_dates(self):
        """처리된 날짜 목록 조회"""
        try:
            # 메타데이터 디렉토리 조회
            response = self.s3_client.list_objects_v2(
                Bucket=self.bucket,
                Prefix=self.dirs['METADATA']
            )
            
            dates = set()  # 중복 제거를 위해 set 사용
            
            # metadata 디렉토리에서 날짜 추출
            for obj in response.get('Contents', []):
                # metadata/YYYY-MM-DD/metadata.json 형식에서 날짜 추출
                parts = obj['Key'].split('/')
                if len(parts) >= 2:
                    date_str = parts[1]
                    if date_str and date_str != "":
                        dates.add(date_str)
            
            # 결과가 없으면 다른 디렉토리도 확인
            if not dates:
                # OCR 결과 디렉토리 확인
                ocr_response = self.s3_client.list_objects_v2(
                    Bucket=self.bucket,
                    Prefix=self.dirs['OCR_RESULTS']
                )
                for obj in ocr_response.get('Contents', []):
                    parts = obj['Key'].split('/')
                    if len(parts) >= 2:
                        date_str = parts[1]
                        if date_str and date_str != "":
                            dates.add(date_str)
                
                # 추출된 PDF 디렉토리 확인
                pdf_response = self.s3_client.list_objects_v2(
                    Bucket=self.bucket,
                    Prefix=self.dirs['EXTRACTED']
                )
                for obj in pdf_response.get('Contents', []):
                    parts = obj['Key'].split('/')
                    if len(parts) >= 2:
                        date_str = parts[1]
                        if date_str and date_str != "":
                            dates.add(date_str)
            
            # 날짜 목록을 리스트로 변환하고 정렬
            dates = sorted(list(dates))
            
            if not dates:
                logger.warning("처리된 날짜를 찾을 수 없습니다.")
                return {"status": "success", "dates": []}
            
            return {"status": "success", "dates": dates}
            
        except Exception as e:
            logger.error(f"처리된 날짜 목록 조회 실패: {e}")
            return {"status": "error", "message": str(e)}

    def save_extracted_pdf(self, date_str, dept_name, page_num, pdf_content):
        """추출된 PDF 저장"""
        try:
            file_key = f"{self.dirs['EXTRACTED']}{date_str}/{dept_name}/page_{page_num}.pdf"
            self.s3_client.put_object(
                Bucket=self.bucket,
                Key=file_key,
                Body=pdf_content
            )
            return {"status": "success", "key": file_key}
        except Exception as e:
            logger.error(f"추출 PDF 저장 실패 ({date_str}/{dept_name}/{page_num}): {e}")
            return {"status": "error", "message": str(e)}

    def get_extracted_pdf(self, date_str, dept_name, page_num):
        """추출된 PDF 조회"""
        try:
            file_key = f"{self.dirs['EXTRACTED']}{date_str}/{dept_name}/page_{page_num}.pdf"
            # S3에서 객체 메타데이터만 확인 (존재 여부만 판단)
            self.s3_client.head_object(Bucket=self.bucket, Key=file_key)
            # 파일이 존재하면 성공 상태와 파일 키 반환
            return {"status": "success", "key": file_key}
        except ClientError as e:
            if e.response['Error']['Code'] == '404' or e.response['Error']['Code'] == 'NoSuchKey':
                return {"status": "not_found"}
            logger.error(f"추출 PDF 조회 실패 ({file_key}): {e}")
            return {"status": "error", "message": str(e)}

    def save_mismatch_data(self, date_str, mismatch_df):
    
        try:
            # 날짜 형식 보정
            if not isinstance(date_str, str) or len(date_str) != 10 or not date_str[:4].isdigit():
                # 판다스, 넘파이, datetime 등 모든 케이스 안전하게 보정
                try:
                    date_str = pd.to_datetime(date_str).strftime('%Y-%m-%d')
                except Exception:
                    date_str = str(date_str)[:10]
            else:
                # 혹시라도 '2025-03-30 00:00:00'처럼 들어온 경우
                if ' ' in date_str:
                    date_str = date_str.split(' ')[0]

            # 저장 전 중복 제거 (같은 날짜 내에서)
            if not mismatch_df.empty:
                before_dedup = len(mismatch_df)
                mismatch_df = mismatch_df.drop_duplicates(subset=['날짜', '부서명', '물품코드'], keep='last')
                after_dedup = len(mismatch_df)
                if before_dedup != after_dedup:
                    logger.info(f"날짜별 저장 시 중복 제거: {before_dedup}개 → {after_dedup}개 (날짜: {date_str})")
                
                # 날짜 형식 보장 (JSON 저장 전)
                if '날짜' in mismatch_df.columns:
                    # 날짜를 문자열로 확실히 변환
                    mismatch_df['날짜'] = mismatch_df['날짜'].astype(str)
                    # YYYY-MM-DD 형식이 아닌 경우 파일명 기준으로 수정
                    invalid_mask = ~mismatch_df['날짜'].str.match(r'^\d{4}-\d{2}-\d{2}$')
                    if invalid_mask.any():
                        logger.warning(f"날짜별 저장 시 잘못된 날짜 형식 {invalid_mask.sum()}개 발견. 파일명({date_str})으로 수정")
                        mismatch_df.loc[invalid_mask, '날짜'] = date_str

            mismatch_key = f"{self.dirs['RESULTS']}{date_str}/mismatches.json"
            json_data = mismatch_df.to_json(orient="records", indent=4, date_format='iso')
            self.s3_client.put_object(
                Bucket=self.bucket,
                Key=mismatch_key,
                Body=json_data
            )
            logger.info(f"불일치 데이터 저장 완료: {mismatch_key}")
            return {"status": "success", "key": mismatch_key}
        except Exception as e:
            logger.error(f"불일치 데이터 저장 실패 ({date_str}): {e}")
            return {"status": "error", "message": str(e)}

    def load_mismatch_data(self, date_str):
    
        try:
            # 날짜 형식 보정(동일)
            if not isinstance(date_str, str) or len(date_str) != 10 or not date_str[:4].isdigit():
                try:
                    date_str = pd.to_datetime(date_str).strftime('%Y-%m-%d')
                except Exception:
                    date_str = str(date_str)[:10]
            else:
                if ' ' in date_str:
                    date_str = date_str.split(' ')[0]

            mismatch_key = f"{self.dirs['RESULTS']}{date_str}/mismatches.json"
            response = self.s3_client.get_object(Bucket=self.bucket, Key=mismatch_key)
            json_data = response['Body'].read()
            df = pd.read_json(io.StringIO(json_data.decode('utf-8')), orient="records")
            logger.info(f"불일치 데이터 로드 완료: {mismatch_key}")
            return {"status": "success", "data": df}
        except ClientError as e:
            if e.response['Error']['Code'] == 'NoSuchKey':
                logger.warning(f"S3에 불일치 데이터 없음 ({mismatch_key})")
                return {"status": "not_found"}
            logger.error(f"불일치 데이터 로드 실패 ({date_str}): {e}")
            return {"status": "error", "message": str(e)}

            
    def save_ocr_text(self, date_str, ocr_text):
        """OCR 텍스트 결과를 S3에 저장"""
        try:
            # 텍스트 파일로 저장 (페이지별)
            for i, page_text in enumerate(ocr_text):
                text_key = f"{self.dirs['OCR_RESULTS']}{date_str}/page_{i+1}.txt"
                self.s3_client.put_object(
                    Bucket=self.bucket,
                    Key=text_key,
                    Body=page_text.encode('utf-8')
                )
            
            # 전체 텍스트 합친 파일 (선택적)
            all_text_key = f"{self.dirs['OCR_RESULTS']}{date_str}/all_pages.txt"
            all_text = "\n\n--- 페이지 구분선 ---\n\n".join(ocr_text)
            self.s3_client.put_object(
                Bucket=self.bucket,
                Key=all_text_key,
                Body=all_text.encode('utf-8')
            )
            
            logger.info(f"OCR 텍스트 저장 완료: {date_str}의 {len(ocr_text)}개 페이지")
            return {"status": "success", "pages": len(ocr_text)}
        except Exception as e:
            logger.error(f"OCR 텍스트 저장 실패 ({date_str}): {e}")
            return {"status": "error", "message": str(e)}
    
    def load_ocr_text(self, date_str):
        """S3에서 OCR 텍스트 결과 로드"""
        try:
            # 먼저 디렉토리 내 모든 텍스트 파일 목록 가져오기
            response = self.s3_client.list_objects_v2(
                Bucket=self.bucket,
                Prefix=f"{self.dirs['OCR_RESULTS']}{date_str}/"
            )
            
            # page_*.txt 형식의 파일만 필터링
            page_files = []
            for obj in response.get('Contents', []):
                if 'page_' in obj['Key'] and obj['Key'].endswith('.txt'):
                    page_num = int(obj['Key'].split('page_')[1].split('.')[0])
                    page_files.append((page_num, obj['Key']))
            
            # 페이지 번호 순으로 정렬
            page_files.sort(key=lambda x: x[0])
            
            ocr_text = []
            for _, page_key in page_files:
                page_response = self.s3_client.get_object(Bucket=self.bucket, Key=page_key)
                page_content = page_response['Body'].read().decode('utf-8')
                ocr_text.append(page_content)
            
            logger.info(f"OCR 텍스트 로드 완료: {date_str}의 {len(ocr_text)}개 페이지")
            return {"status": "success", "data": ocr_text}
        except ClientError as e:
            if e.response['Error']['Code'] == 'NoSuchKey':
                logger.warning(f"S3에 OCR 텍스트 없음 ({date_str})")
                return {"status": "not_found"}
            logger.error(f"OCR 텍스트 로드 실패 ({date_str}): {e}")
            return {"status": "error", "message": str(e)}
    
    def get_file_hash(self, file_obj):
        """파일 내용의 MD5 해시값 계산"""
        try:
            file_obj.seek(0)
            file_bytes = file_obj.read()
            file_hash = hashlib.md5(file_bytes).hexdigest()
            file_obj.seek(0)  # 파일 포인터 초기화
            return {"status": "success", "hash": file_hash}
        except Exception as e:
            logger.error(f"파일 해시 계산 실패: {e}")
            return {"status": "error", "message": str(e)}
    
    def check_file_exists(self, date_str, file_hash, file_type):
        """해시값으로 동일한 파일이 이미 S3에 존재하는지 확인"""
        try:
            # 해시 정보가 저장된 메타데이터 조회
            metadata_result = self.load_metadata(date_str)
            if metadata_result["status"] == "success":
                metadata = metadata_result["data"]
                
                # 파일 타입에 따라 확인
                if file_type == "PDF" and "pdf_hash" in metadata:
                    # 해시값 비교
                    if metadata["pdf_hash"] == file_hash:
                        return {
                            "status": "success", 
                            "exists": True, 
                            "metadata": metadata
                        }
                
                elif file_type == "EXCEL" and "excel_hash" in metadata:
                    if metadata["excel_hash"] == file_hash:
                        return {
                            "status": "success", 
                            "exists": True, 
                            "metadata": metadata
                        }
            
            # 파일이 없거나 해시값이 다름
            return {"status": "success", "exists": False}
        
        except Exception as e:
            logger.error(f"파일 존재 여부 확인 실패 ({date_str}, {file_type}): {e}")
            return {"status": "error", "message": str(e)}

    def save_missing_items_by_date(self, missing_df, date_str):


        try:
            mismatch_key = f"{self.dirs['RESULTS']}{date_str}/mismatches.json"
            try:
                s3_obj = self.s3_client.get_object(Bucket=self.bucket, Key=mismatch_key)
                json_bytes = s3_obj["Body"].read()
                mismatch_df = pd.read_json(io.BytesIO(json_bytes), orient="records")
            except Exception:
                # 기존 파일이 없으면 빈 DF로 시작
                mismatch_df = pd.DataFrame()

            # 1. 컬럼 구조 맞추기 - 표준 컬럼 순서 정의
            standard_columns = ['날짜', '부서명', '물품코드', '물품명', '청구량', '수령량', '차이', '누락']
            
            # 기존 데이터가 없으면 표준 컬럼으로 초기화
            if mismatch_df.empty:
                mismatch_df = pd.DataFrame(columns=standard_columns)
            
            # missing_df에 누락된 컬럼 추가
            for col in standard_columns:
                if col not in missing_df.columns:
                    missing_df[col] = ""
                if col not in mismatch_df.columns:
                    mismatch_df[col] = ""
            
            # 표준 컬럼 순서로 정렬
            missing_df = missing_df[standard_columns]
            mismatch_df = mismatch_df[standard_columns]

            # 2. 기존 + 신규(전산누락) append/concat
            combined = pd.concat([mismatch_df, missing_df], ignore_index=True)

            # 3. 중복제거 시 누락 컬럼 보존 로직 추가
            before_dedup = len(combined)
            # 날짜, 부서명, 물품코드가 같은 항목들을 그룹화하여 누락 컬럼 병합
            if not combined.empty:
                # 중복 항목들을 찾아서 누락 컬럼 병합
                duplicated_mask = combined.duplicated(subset=['날짜', '부서명', '물품코드'], keep=False)
                if duplicated_mask.any():
                    # 중복된 항목들 처리
                    duplicated_items = combined[duplicated_mask].copy()
                    non_duplicated_items = combined[~duplicated_mask].copy()
                    
                    # 중복 항목들을 그룹화하여 누락 컬럼 병합
                    merged_duplicates = []
                    for (date, dept, code), group in duplicated_items.groupby(['날짜', '부서명', '물품코드']):
                        # 가장 최근 항목을 기본으로 사용
                        merged_item = group.iloc[-1].copy()
                        
                        # 누락 컬럼 병합: null이 아닌 값 우선 사용
                        missing_values = group['누락'].dropna()
                        if not missing_values.empty:
                            # 전산누락이 있으면 우선 사용, 없으면 다른 값 사용
                            if '전산누락' in missing_values.values:
                                merged_item['누락'] = '전산누락'
                            else:
                                merged_item['누락'] = missing_values.iloc[-1]
                        
                        merged_duplicates.append(merged_item)
                    
                    # 중복 제거된 항목들과 비중복 항목들 결합
                    if merged_duplicates:
                        combined = pd.concat([non_duplicated_items, pd.DataFrame(merged_duplicates)], ignore_index=True)
                    else:
                        combined = non_duplicated_items
                else:
                    # 중복이 없으면 그대로 사용
                    pass
            
            after_dedup = len(combined)
            logger.info(f"중복 제거 (누락 컬럼 보존): {before_dedup}개 → {after_dedup}개")

            # 4. 날짜별 파일 저장 (통합 작업은 부서별 통계 탭에서 수동 실행)
            mismatch_json = combined.to_json(orient="records", indent=4)
            self.s3_client.put_object(Bucket=self.bucket, Key=mismatch_key, Body=mismatch_json)
            logger.info(f"날짜별 mismatches.json({date_str}) 저장/업데이트 완료: {len(combined)}개")
            
            # 저장 직후 확인 (디버깅용)
            try:
                verify_result = self.s3_client.get_object(Bucket=self.bucket, Key=mismatch_key)
                verify_data = json.loads(verify_result['Body'].read())
                logger.info(f"저장 확인: {mismatch_key}에 {len(verify_data)}개 항목 존재")
                # 전산누락 항목 확인
                missing_count = sum(1 for item in verify_data if '누락' in item and '누락' in str(item.get('누락', '')))
                logger.info(f"저장 확인: 전산누락 항목 {missing_count}개 포함")
            except Exception as e:
                logger.error(f"저장 확인 실패: {e}")

            
            return {"status": "success", "message": f"{date_str} 전산누락 데이터 저장 완료. 부서별 통계에서 '날짜별 작업 내용 병합' 버튼을 눌러주세요."}

        except Exception as e:
            logger.error(f"날짜별 mismatches 저장 실패: {e}", exc_info=True)
            return {"status": "error", "message": str(e)}
    
    def list_all_dates_in_results(self):

        prefix = self.dirs['RESULTS']
        paginator = self.s3_client.get_paginator('list_objects_v2')
        operation_parameters = {'Bucket': self.bucket, 'Prefix': prefix, 'Delimiter': '/'}
        page_iterator = paginator.paginate(**operation_parameters)

        date_folders = set()
        date_pattern = re.compile(r'(\d{4}-\d{2}-\d{2})/')  # 2025-05-21/ 패턴

        for page in page_iterator:
            if "CommonPrefixes" in page:
                for cp in page["CommonPrefixes"]:
                    folder = cp["Prefix"][len(prefix):]
                    match = date_pattern.match(folder)
                    if match:
                        date_folders.add(match.group(1))
            # (혹시 날짜 폴더가 Prefix 말고 Key에서만 발견되는 구조라면 아래 코드도 추가)
            if "Contents" in page:
                for obj in page["Contents"]:
                    key = obj["Key"][len(prefix):]
                    parts = key.split('/')
                    if len(parts) > 1 and re.match(r'\d{4}-\d{2}-\d{2}', parts[0]):
                        date_folders.add(parts[0])

        return sorted(list(date_folders))

    def update_full_mismatches_json(self):
        """날짜별 mismatches.json 파일들을 통합하여 전체 파일 생성"""
        try:
            prefix = f"{self.dirs['RESULTS']}"
            date_folders = self.list_all_dates_in_results()

            if not date_folders:
                logger.warning("통합할 날짜별 mismatches.json 파일이 없습니다.")
                return {"status": "error", "message": "날짜별 파일 없음"}

            all_mismatches = []
            for date_str in date_folders:
                key = f"{prefix}{date_str}/mismatches.json"
                try:
                    s3_obj = self.s3_client.get_object(Bucket=self.bucket, Key=key)
                    json_bytes = s3_obj["Body"].read()
                    df = pd.read_json(io.BytesIO(json_bytes), orient="records")

                    # 날짜 컬럼 표준화
                    if '날짜' in df.columns:
                        df['날짜'] = pd.to_datetime(df['날짜'], errors='coerce').dt.strftime('%Y-%m-%d')
                        invalid_dates = df['날짜'].isna()
                        if invalid_dates.any():
                            logger.warning(f"{date_str}: 날짜 형식 오류 발견. 파일명 기준으로 수정")
                            df.loc[invalid_dates, '날짜'] = date_str
                    else:
                        df['날짜'] = date_str

                    if '누락' in df.columns:
                        missing_items = df[df['누락'].str.contains('누락', na=False)]
                        normal_items = df[~df['누락'].str.contains('누락', na=False)]

                        # 전산누락 항목만 있는 경우
                        if not missing_items.empty and normal_items.empty:
                            logger.info(f"{date_str}: 전산누락 항목만 존재 ({len(missing_items)}개)")
                            all_mismatches.append(missing_items)

                        # 일반 불일치 항목만 있는 경우
                        elif missing_items.empty and not normal_items.empty:
                            logger.info(f"{date_str}: 일반 불일치 항목만 존재 ({len(normal_items)}개)")
                            all_mismatches.append(normal_items)

                        # 둘 다 있는 경우
                        elif not missing_items.empty and not normal_items.empty:
                            logger.info(f"{date_str}: 일반 불일치({len(normal_items)}개), 전산누락({len(missing_items)}개) 존재")
                            combined_df = pd.concat([normal_items, missing_items], ignore_index=True)
                            all_mismatches.append(combined_df)

                        # 둘 다 없는 경우는 건너뜀
                    else:
                        # 누락 컬럼 자체가 없을 경우 전체 추가
                        logger.info(f"{date_str}: '누락' 컬럼 없음, 전체 추가")
                        all_mismatches.append(df)

                except Exception as e:
                    logger.warning(f"{key} 로드 실패: {e}")
                    continue
    
            if not all_mismatches:
                logger.warning("유효한 mismatches 데이터가 없습니다.")
                return {"status": "error", "message": "유효한 데이터 없음"}

            # 데이터 통합
            merged_df = pd.concat(all_mismatches, ignore_index=True)
            logger.info(f"날짜별 파일 통합 후 총 항목 수: {len(merged_df)}개")

            # 중복 제거
            before_dedup = len(merged_df)
            merged_df = merged_df.drop_duplicates(subset=['날짜', '부서명', '물품코드'], keep='last')
            after_dedup = len(merged_df)
            logger.info(f"중복 제거: {before_dedup}개 → {after_dedup}개")

            # 완료 처리 필터링 적용
            try:
                completion_logs_result = self.load_completion_logs()
                if completion_logs_result["status"] == "success":
                    completion_logs = completion_logs_result["data"]
                    if completion_logs:
                        missing_mask = merged_df['누락'].str.contains('누락', na=False) if '누락' in merged_df.columns else pd.Series([False] * len(merged_df))
                        missing_items = merged_df[missing_mask].copy()
                        regular_items = merged_df[~missing_mask].copy()
    
                        if not regular_items.empty:
                            regular_items = filter_completed_items(regular_items, completion_logs)

                        merged_df = pd.concat([regular_items, missing_items], ignore_index=True)
                        logger.info("완료 처리 필터링 완료")
                    else:
                        logger.info("완료 처리 로그가 없어 필터링을 건너뜁니다.")
                else:
                    logger.warning("완료 처리 로그 로드 실패. 필터링 없이 진행합니다.")
            except Exception as filter_err:
                logger.error(f"완료 처리 필터링 중 오류: {filter_err}. 필터링 없이 진행합니다.")

            # 데이터 정렬
            merged_df = merged_df.sort_values(['날짜', '부서명', '물품코드'])

            # 저장
            full_mismatches_key = f"{self.dirs['RESULTS']}mismatches_full.json"
            json_data = merged_df.to_json(orient="records", indent=4)
            self.s3_client.put_object(
                Bucket=self.bucket,
                Key=full_mismatches_key,
                Body=json_data
            )

            logger.info(f"전체 통합 mismatches_full.json 저장 완료: {len(merged_df)}개 항목")
            return {"status": "success", "count": len(merged_df)}

        except Exception as e:
            logger.error(f"전체 통합 mismatches_full.json 생성 중 오류: {e}")
            return {"status": "error", "message": str(e)}



    def load_full_mismatches(self):
      
        full_mismatches_key = f"{self.dirs['RESULTS']}mismatches_full.json"
        try:
            s3_obj = self.s3_client.get_object(Bucket=self.bucket, Key=full_mismatches_key)
            json_bytes = s3_obj["Body"].read()
            df = pd.read_json(io.BytesIO(json_bytes), orient="records")
            return df
        except Exception as e:
            logger.error(f"전체 통합 mismatches_full.json 불러오기 실패: {e}")
            return pd.DataFrame()


    def save_pdf_preview_image(self, date_str, dept_name, page_num, img_obj: Image.Image):
        """
    PDF 미리보기 이미지를 썸네일로 변환해 S3에 저장하고,
    날짜별 메타데이터(preview_images)에 정보 반영.
    """
        try:
            if self.s3_client is None:
                logger.error("S3 클라이언트가 초기화되지 않았습니다.")
                return {"status": "error", "message": "S3 연결 실패"}

            # 부서명 폴더/파일명 안전화
            safe_dept_name = dept_name.replace('/', '_').replace('\\', '_')

            # --- 썸네일 변환 (예: 400x600) ---
            img = img_obj.copy()
            img.thumbnail((700, 1000))  # 비율유지 최대 400x600

            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG', optimize=True, compress_level=6)
            img_byte_arr.seek(0)

            # --- 파일 경로 ---
            file_key = f"preview_images/{date_str}/{safe_dept_name}_page{page_num}_preview.png"
            
            try:
                self.s3_client.upload_fileobj(img_byte_arr, self.bucket, file_key)
            except ClientError as ce:
                logger.error(f"S3 업로드 실패: {ce}")
                return {"status": "error", "message": f"S3 업로드 실패: {ce.response.get('Error',{}).get('Message', '알 수 없음')}"}

            # --- 메타데이터에 이미지 정보 반영 ---
            metadata_result = self.load_metadata(date_str)
            if metadata_result.get("status") != "success":
                metadata = {}
            else:
                metadata = metadata_result.get("data", {})
            if not isinstance(metadata, dict):
                metadata = {}

            if "preview_images" not in metadata:
                metadata["preview_images"] = []

            image_info = {
                "dept": dept_name,  # 원본 부서명(표시용)
                "page": page_num,
                "file_key": file_key
            }

            # 기존에 동일 부서/페이지가 있으면 업데이트
            exists = False
            for img_info_item in metadata["preview_images"]:
                if img_info_item.get("dept") == dept_name and img_info_item.get("page") == page_num:
                    img_info_item["file_key"] = file_key
                    exists = True
                    break
            if not exists:
                metadata["preview_images"].append(image_info)

            save_result = self.save_metadata(date_str, metadata)
            if save_result.get("status") != "success":
                logger.warning(f"메타데이터 저장 실패 ({date_str}): {save_result.get('message')}")

            return {"status": "success", "message": "썸네일 이미지 저장 및 메타데이터 기록 완료", "file_key": file_key}

        except Exception as e:
            logger.error(f"이미지 저장 중 예외 발생: {e}", exc_info=True)
            return {"status": "error", "message": f"이미지 저장 오류: {str(e)}"}
        


    def save_completion_log(self, completed_items):
        """완료 처리 로그를 JSON 형태로 S3에 저장 (강화된 유효성 검사)"""
        try:
            log_key = f"{self.dirs['RESULTS']}completion_logs.json"
            logger.info(f"완료 처리 로그 저장 시작 - 입력 항목 수: {len(completed_items) if isinstance(completed_items, list) else 'None (입력값이 리스트가 아님)'}")

            # 입력 데이터 검증 (리스트 여부)
            if not isinstance(completed_items, list):
                logger.error(f"completed_items는 리스트여야 합니다. 실제 타입: {type(completed_items)}")
                return {"status": "error", "message": "잘못된 데이터 형식 (리스트가 아님)"}

            # numpy/pandas 타입을 파이썬 기본 타입으로 변환하는 함수
            def convert_to_serializable(item):
                if not isinstance(item, dict):
                    # 이 경우는 아래에서 처리하므로 여기서는 None 반환
                    return None
                converted = {}
                for key, value in item.items():
                    if hasattr(value, 'item'):  # numpy type check
                        converted[key] = value.item()
                    elif pd.isna(value):  # null check
                        converted[key] = None
                    else:
                        converted[key] = value
                return converted

            # 기존 로그 파일이 있으면 로드
            existing_logs = []
            try:
                response = self.s3_client.get_object(Bucket=self.bucket, Key=log_key)
                loaded_content = response['Body'].read().decode('utf-8')
                if loaded_content: # 파일 내용이 비어있지 않은 경우에만 파싱 시도
                    existing_logs = json.loads(loaded_content)
                    if not isinstance(existing_logs, list):
                        logger.warning(f"기존 로그 파일({log_key})이 리스트가 아닙니다 (타입: {type(existing_logs)}). 새로 초기화합니다.")
                        existing_logs = []
                    else:
                        # 기존 로그도 검증 (딕셔너리, 필수 키)
                        valid_existing_logs = []
                        for i, log_item in enumerate(existing_logs):
                            if not isinstance(log_item, dict):
                                logger.warning(f"기존 로그 {i}번째 항목이 딕셔너리가 아님: {log_item}")
                                continue
                            if not all(k in log_item for k in ['날짜', '부서명', '물품코드']): # 물품코드도 필수적으로 검사
                                logger.warning(f"기존 로그 {i}번째 항목에 필수 키(날짜, 부서명, 물품코드) 누락: {log_item}")
                                continue
                            valid_existing_logs.append(log_item)
                        existing_logs = valid_existing_logs
                        logger.info(f"기존 로그 파일 로드 및 검증 완료 - 유효 항목 수: {len(existing_logs)}")
                else:
                    logger.info(f"기존 로그 파일({log_key})이 비어있습니다. 새로 생성합니다.")
            except ClientError as e:
                if e.response['Error']['Code'] == 'NoSuchKey':
                    logger.info(f"기존 로그 파일({log_key})이 없어 새로 생성합니다.")
                else:
                    logger.warning(f"기존 로그 파일 로드 중 S3 오류 발생({log_key}): {e}")
            except json.JSONDecodeError as e:
                logger.warning(f"기존 로그 파일 파싱 중 JSON 오류 발생({log_key}): {e}")
            except Exception as e:
                logger.warning(f"기존 로그 파일 처리 중 예상치 못한 오류({log_key}): {e}")

            # 중복 제거를 위한 키 생성 함수
            def get_item_key(item):
                # 이 함수는 item이 dict라고 가정하고 호출됨
                return f"{item.get('날짜', '')}_{item.get('부서명', '')}_{item.get('물품코드', '')}"

            # 기존 로그에서 키 집합 생성 (이미 검증된 로그 사용)
            existing_keys = {get_item_key(item) for item in existing_logs}

            new_items_to_add = []
            invalid_item_count = 0

            for i, item in enumerate(completed_items):
                # 1. 각 항목이 딕셔너리인지 확인
                if not isinstance(item, dict):
                    logger.warning(f"새로 추가할 {i}번째 항목이 딕셔너리가 아님: {item} (타입: {type(item)}). 건너뜁니다.")
                    invalid_item_count += 1
                    continue

                # 2. 필수 키 확인 (날짜, 부서명, 물품코드) - 물품코드 누락 시에도 건너뜀
                required_keys = ['날짜', '부서명', '물품코드']
                missing_keys = [key for key in required_keys if key not in item or pd.isna(item[key])] # NaN/None도 누락으로 간주
                if missing_keys:
                    logger.warning(f"새로 추가할 {i}번째 항목에 필수 키 {missing_keys} 누락 또는 값 없음: {item}. 건너뜁니다.")
                    invalid_item_count += 1
                    continue

                # 3. 데이터 타입 변환 (원본 item에 대해)
                serializable_item = convert_to_serializable(item) # convert_to_serializable은 이미 item이 dict임을 가정
                if not serializable_item: # 변환 실패 (내부 로직상 거의 발생 안함)
                    logger.warning(f"새로 추가할 {i}번째 항목 직렬화 실패: {item}. 건너뜁니다.")
                    invalid_item_count += 1
                    continue
                
                # 날짜 형식 표준화 (YYYY-MM-DD)
                try:
                    serializable_item['날짜'] = pd.to_datetime(serializable_item['날짜']).strftime('%Y-%m-%d')
                except Exception as e:
                    logger.warning(f"새로 추가할 {i}번째 항목의 날짜 형식 변환 실패 ('{serializable_item.get('날짜')}'): {e}. 건너뜁니다.")
                    invalid_item_count += 1
                    continue

                # 4. 중복 확인 (변환된 serializable_item 기준)
                item_key = get_item_key(serializable_item) # 여기서 serializable_item은 항상 dict
                if item_key not in existing_keys:
                    new_items_to_add.append(serializable_item)
                    existing_keys.add(item_key)
                # else:
                #     logger.info(f"새로 추가할 {i}번째 항목은 이미 존재: {item_key}") # 중복은 로그 안 남김 (너무 많을 수 있음)

            logger.info(f"새로 추가될 항목 수: {len(new_items_to_add)}, 유효하지 않아 건너뛴 항목 수: {invalid_item_count}")

            if not new_items_to_add and invalid_item_count == len(completed_items) and len(completed_items) > 0:
                logger.warning("모든 입력 항목이 유효하지 않아 추가할 새 로그가 없습니다.")
                return {"status": "no_valid_items", "message": "유효한 로그 항목 없음", "added_items": 0}

            if new_items_to_add:
                all_logs_to_save = existing_logs + new_items_to_add
                try:
                    # JSON으로 변환하여 저장
                    json_data = json.dumps(all_logs_to_save, ensure_ascii=False, indent=2)
                    self.s3_client.put_object(
                        Bucket=self.bucket,
                        Key=log_key,
                        Body=json_data.encode('utf-8')
                    )
                    logger.info(f"완료 처리 로그 저장 성공 ({log_key}) - 총 {len(all_logs_to_save)}개 항목 저장 (새 항목 {len(new_items_to_add)}개 추가).")
                    return {"status": "success", "key": log_key, "added_items": len(new_items_to_add), "total_items": len(all_logs_to_save)}
                except Exception as e:
                    logger.error(f"S3 업로드 중 오류 발생({log_key}): {e}")
                    return {"status": "error", "message": f"S3 업로드 실패: {str(e)}"}
            else:
                logger.info(f"추가할 새로운 유효 항목이 없습니다 (기존 로그 수: {len(existing_logs)}). 저장 작업 건너뜁니다.")
                return {"status": "success", "key": log_key, "added_items": 0, "total_items": len(existing_logs), "message": "새로 추가된 항목 없음"}

        except Exception as e:
            logger.error(f"완료 처리 로그 저장 중 예상치 못한 최상위 오류 발생: {e}", exc_info=True)
            return {"status": "error", "message": f"예상치 못한 오류: {str(e)}"}

    def load_completion_logs(self):
        """완료 처리 로그를 S3에서 로드 (강화된 유효성 검사)"""
        try:
            log_key = f"{self.dirs['RESULTS']}completion_logs.json"
            
            # S3에서 파일 가져오기
            try:
                response = self.s3_client.get_object(Bucket=self.bucket, Key=log_key)
                file_content = response['Body'].read().decode('utf-8')
            except ClientError as e:
                if e.response['Error']['Code'] == 'NoSuchKey':
                    logger.info(f"completion_logs.json 파일({log_key})이 없습니다.")
                    return {"status": "not_found", "data": []}
                logger.error(f"완료 처리 로그 S3에서 로드 실패 ({log_key}): {e}")
                return {"status": "error", "data": [], "message": f"S3 로드 오류: {str(e)}"}
            except Exception as e:
                logger.error(f"완료 처리 로그 S3에서 읽는 중 예외 발생 ({log_key}): {e}", exc_info=True)
                return {"status": "error", "data": [], "message": f"S3 파일 읽기 오류: {str(e)}"}

            # 파일 내용이 비어있는 경우
            if not file_content:
                logger.info(f"completion_logs.json 파일({log_key}) 내용은 비어있습니다.")
                return {"status": "success", "data": []} # 빈 파일도 성공으로 간주, 빈 리스트 반환

            # JSON 파싱
            try:
                logs = json.loads(file_content)
            except json.JSONDecodeError as e:
                logger.error(f"completion_logs.json 파일({log_key}) JSON 파싱 오류: {e}. 파일 내용 일부: {file_content[:200]}")
                return {"status": "error", "data": [], "message": f"잘못된 JSON 형식: {str(e)}"}
            
            # 데이터 검증 및 정제 (리스트 타입 확인)
            if not isinstance(logs, list):
                logger.warning(f"completion_logs.json에 리스트 이외의 자료가 있음 (타입: {type(logs)}). 파일 내용: {file_content[:200]}")
                return {"status": "error", "data": [], "message": "저장된 로그가 리스트가 아님"}
                
            # 각 항목 검증 (딕셔너리, 필수 키, 날짜 형식)
            valid_logs = []
            invalid_count = 0
            for i, item in enumerate(logs):
                if not isinstance(item, dict):
                    logger.warning(f"로그 항목 {i}가 딕셔너리가 아님: {item}")
                    invalid_count += 1
                    continue
                    
                required_keys = ['날짜', '부서명', '물품코드'] # 물품코드도 필수로 검사
                missing_keys = [key for key in required_keys if key not in item or pd.isna(item[key])]
                if missing_keys:
                    logger.warning(f"로그 항목 {i}에 필수 키 {missing_keys} 누락 또는 값 없음: {item}")
                    invalid_count += 1
                    continue
                    
                # 날짜 형식 검증 및 표준화 (YYYY-MM-DD)
                try:
                    # 표준화 시도. 실패하면 건너뜀
                    item['날짜'] = pd.to_datetime(item['날짜']).strftime('%Y-%m-%d')
                except Exception as e:
                    logger.warning(f"로그 항목 {i}의 날짜 형식 ('{item.get('날짜')}') 변환 실패: {e}. 건너뜁니다.")
                    invalid_count += 1
                    continue
                    
                valid_logs.append(item)
            
            if invalid_count > 0:
                logger.warning(f"총 {len(logs)}개 로그 중 {invalid_count}개 항목이 유효하지 않아 제외되었습니다.")
            
            logger.info(f"유효한 완료 로그 {len(valid_logs)}개 로드 완료 ({log_key}).")
            return {"status": "success", "data": valid_logs}
            
        except Exception as e:
            logger.error(f"완료 처리 로그 로드 중 예상치 못한 최상위 오류 발생: {e}", exc_info=True)
            return {"status": "error", "data": [], "message": f"예상치 못한 오류: {str(e)}"}


# --- 날짜 표준화 함수 (streamlit_app.py 내에 직접 정의) ---
def standardize_date(date_str):
    """다양한 형식의 날짜 문자열을 YYYY-MM-DD로 표준화합니다.
    
    입력:
    - date_str: 날짜 문자열 (파일명, 시트명 등)
    
    처리하는 형식:
    1. YYYY-MM-DD 또는 YYYY.MM.DD
    2. MM.DD, MM-DD, M.D, M-D
    3. 파일명에서 날짜 패턴 추출 (예: "불출대장_12.15.pdf" -> "12.15")
    
    출력:
    - 표준화된 날짜 (YYYY-MM-DD 형식)
    - 날짜가 아닌 경우 원본 반환
    """
    now = datetime.now()
    year = now.year  # 기본 연도는 현재 연도
    
    # 파일명에서 날짜 패턴 추출 시도
    # 파일명에서 MM.DD 패턴 추출
    file_date_match = re.search(r'(\d{1,2})[.-](\d{1,2})', str(date_str))
    if file_date_match:
        try:
            m, d = map(int, file_date_match.groups())
            # 실제 존재하는 날짜인지 검증
            if 1 <= m <= 12 and 1 <= d <= 31:
                return datetime(year, m, d).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # YYYY-MM-DD 또는 YYYY.MM.DD 형식 확인
    match_ymd = re.match(r'(\d{4})[.-]?(\d{1,2})[.-]?(\d{1,2})', str(date_str).strip())
    if match_ymd:
        try:
            y, m, d = map(int, match_ymd.groups())
            return datetime(y, m, d).strftime('%Y-%m-%d')
        except ValueError:
            pass  # 잘못된 날짜면 다음 패턴 시도

    # MM.DD, MM-DD, M.D, M-D 형식 확인 (마침표 포함)
    match_md = re.match(r'(\d{1,2})[.-](\d{1,2})\.?$', str(date_str).strip())
    if match_md:
        try:
            m, d = map(int, match_md.groups())
            # 연도 추정 - 현재보다 미래 날짜면 작년으로 처리
            date_with_current_year = datetime(year, m, d)
            if date_with_current_year > now and m > now.month:
                year -= 1
            return datetime(year, m, d).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # 날짜 형식을 인식할 수 없는 경우 원본 반환
    logger.warning(f"날짜 형식 인식 불가: {date_str}")
    return str(date_str).strip()  # 입력값을 문자열로 반환
# ----------------------------------------------------

# --- 완료 처리 항목 필터링 유틸리티 함수 ---
def is_item_completed(item, completion_logs):
    """주어진 항목이 완료 처리 로그에 있는지 확인합니다.
    
    Args:
        item (dict): 불일치 데이터 항목 (날짜, 부서명, 물품코드 포함)
        completion_logs (list): 완료 처리 로그 목록
    
    Returns:
        bool: 완료 처리 여부
    """
    for log in completion_logs:
        if (str(log.get('날짜')) == str(item.get('날짜')) and
            str(log.get('부서명')) == str(item.get('부서명')) and
            str(log.get('물품코드')) == str(item.get('물품코드'))):
            return True
    return False

def filter_completed_items(mismatch_data, completion_logs, date_range=None):
    """완료 처리된 항목을 필터링하는 함수
    
    Args:
        mismatch_data: 불일치 데이터 DataFrame
        completion_logs: 완료 처리 로그 리스트
        date_range: (start_date, end_date) 튜플, None이면 전체 기간
    """
    try:
        if mismatch_data.empty or not completion_logs:
            return mismatch_data

        filtered_completion_logs = completion_logs
        if date_range:
            start_date, end_date = date_range
            filtered_completion_logs = []
            for log in completion_logs:
                try:
                    log_date = pd.to_datetime(log.get('날짜', '')).date()
                    if start_date <= log_date <= end_date:
                        filtered_completion_logs.append(log)
                except:
                    continue

        completed_items = set()
        invalid_completion_logs = 0
        for log in filtered_completion_logs:
            try:
                date = str(log.get('날짜', ''))
                dept = str(log.get('부서명', ''))
                code = str(log.get('물품코드', ''))
                if date and dept and code:
                    date = pd.to_datetime(date).strftime('%Y-%m-%d')
                    completed_key = f"{date}_{dept}_{code}"
                    completed_items.add(completed_key)
                else:
                    invalid_completion_logs += 1
            except:
                invalid_completion_logs += 1

        missing_mask = mismatch_data['누락'].str.contains('누락', na=False) if '누락' in mismatch_data.columns else pd.Series([False] * len(mismatch_data))
        missing_items = mismatch_data[missing_mask].copy()
        regular_items = mismatch_data[~missing_mask].copy()

        if not regular_items.empty:
            regular_items = regular_items[
                ~regular_items.apply(lambda row: f"{row['날짜']}_{row['부서명']}_{row['물품코드']}", axis=1).isin(completed_items)
            ]

        filtered_data = pd.concat([regular_items, missing_items], ignore_index=True)

        return filtered_data

    except Exception as e:
        logger.error(f"완료 항목 필터링 중 오류 발생: {e}", exc_info=True)
        return mismatch_data


# 삭제된 중복 함수
# ----------------------------------------------------

# 한글 폰트 설정 함수
def set_korean_font():
    try:
        # 시스템 폰트 검색
        font_path = None
        font_files = fm.findSystemFonts(fontpaths=None, fontext='ttf')
        
        # Windows: Malgun Gothic
        if os.name == 'nt':
            for fpath in font_files:
                if 'malgun' in fpath.lower():
                    font_path = fpath
                    break
        # macOS: AppleGothic
        elif os.name == 'posix':
            for fpath in font_files:
                if 'applegothic' in fpath.lower():
                    font_path = fpath
                    break
        # Linux: NanumGothic (설치 필요)
        else:
            for fpath in font_files:
                if 'nanumgothic' in fpath.lower():
                    font_path = fpath
                    break

        if font_path:
            plt.rc('font', family=fm.FontProperties(fname=font_path).get_name())
            plt.rcParams['axes.unicode_minus'] = False # 마이너스 기호 깨짐 방지
        else:
            logger.warning("적절한 한글 폰트를 찾지 못했습니다. 시스템에 폰트 설치를 권장합니다.")
            # 기본 폰트로 진행
            plt.rc('font', family='sans-serif')
            plt.rcParams['axes.unicode_minus'] = False
            
    except Exception as e:
        logger.error(f"한글 폰트 설정 중 오류 발생: {e}")
        # 오류 발생 시 기본 폰트로 진행
        plt.rc('font', family='sans-serif')
        plt.rcParams['axes.unicode_minus'] = False


# 앱 스타일
st.markdown("""
<style>
    /* 콘텐츠 영역 기본 텍스트 크기 */
    .main .block-container div[data-testid=\"stMarkdownContainer\"],
    .main .block-container div[data-testid=\"stText\"] {
        font-size: 18px !important;
    }
    /* 버튼 폰트 크기 */
    button {
        font-size: 17px !important; /* 버튼은 약간 작게 */
    }
    /* st.dataframe 내부 테이블 폰트 크기 */
    .stDataFrame table th,
    .stDataFrame table td {
        font-size: 18px !important; /* 다른 콘텐츠와 동일하게 */
    }
    h1, h2, h3, h4, h5, h6 {
        /* 헤더 폰트 크기는 기본값을 유지하거나 필요시 별도 조정 */
    }
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .stProgress .st-eb {
        background-color: #4CAF50;
    }
    .stTabs [data-baseweb=\"tab-list\"] {
        gap: 1px;
    }
    .stTabs [data-baseweb=\"tab\"] {
        height: 50px;
        white-space: normal !important; /* 자동 줄바꿈 허용 */
        background-color: #F0F2F6;
        border-radius: 4px 4px 0 0;
        padding-left: 1rem;
        padding-right: 1rem;
        /* 탭 제목 크기 명시적 설정 */
        font-size: 17px !important; /* 탭 제목 크기 유지 */
        min-width: 100px !important; /* 탭의 최소 너비 조정 (필요한 너비로 설정 가능) */
        word-break: keep-all; /* 단어 단위 줄바꿈 활성화 (한글 기준) */
        padding: 0.5rem !important; /* 적당한 내부 패딩 추가 */
        height: auto !important; /* 탭 높이 자동 조정 */
    }
    .stTabs [aria-selected=\"true\"] {
        background-color: #E0E0E0;
    }
    .department-card {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #f8f9fa;
        margin-bottom: 0.5rem;
    }
    .mismatch-badge {
        background-color: #ff4b4b;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 1rem;
        font-size: 0.8rem;
        margin-left: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# 세션 상태 초기화 (수정 및 추가)
if 'ocr_results_by_date' not in st.session_state:
    st.session_state.ocr_results_by_date = {} # 날짜별 OCR 결과 저장
if 'pdf_paths_by_date' not in st.session_state:
    st.session_state.pdf_paths_by_date = {} # 날짜별 원본 PDF 경로 저장
if 'processed_pdfs_by_date' not in st.session_state:
    st.session_state.processed_pdfs_by_date = {} # 날짜별 처리된 PDF 경로 저장 (fitz 객체 대신 경로 저장 권장)
if 'dept_page_tuples_by_date' not in st.session_state:
    st.session_state.dept_page_tuples_by_date = {} # 날짜별 부서-페이지 튜플 목록 저장

if 'excel_dates' not in st.session_state:
    st.session_state.excel_dates = [] # 원본 엑셀 날짜 (시트명) - 현재 사용 안함
if 'standardized_excel_dates' not in st.session_state:
    st.session_state.standardized_excel_dates = [] # 표준화된 엑셀 날짜
if 'pdf_dates' not in st.session_state:
    st.session_state.pdf_dates = []        # 표준화된 PDF 날짜
if 'available_dates' not in st.session_state:
     st.session_state.available_dates = [] # 통합 날짜 목록
if 'selected_date' not in st.session_state:
    st.session_state.selected_date = None
if 'item_db' not in st.session_state:
    st.session_state.item_db = {}  # 물품 코드-이름 매핑 DB
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = pd.DataFrame()  # 엑셀 데이터
if 'mismatch_data' not in st.session_state:
    st.session_state.mismatch_data = pd.DataFrame()  # 불일치 데이터
if 'missing_items' not in st.session_state:
    st.session_state.missing_items = pd.DataFrame()  # 누락 품목 데이터
if 'receipt_status' not in st.session_state:
    st.session_state.receipt_status = {} # 날짜-부서별 인수증 상태 저장 ('인수증 없음')
if 'missing_receipt_info' not in st.session_state:
    st.session_state.missing_receipt_info = {}  # 부서별 날짜 목록을 저장할 딕셔너리
# 완료 처리된 항목을 세션에 저장하는 변수 추가
if 'completion_logs' not in st.session_state:
    st.session_state.completion_logs = []  # 완료 처리 로그


# PDF에서 이미지 추출 함수
@st.cache_data(ttl=3600, max_entries=100)
def extract_pdf_preview(pdf_path_or_bytes, page_num=0, dpi=120, thumbnail_size=(700, 1000)):
    """
    PDF 파일의 특정 페이지를 썸네일(미리보기) 이미지로 추출하여 반환 (PIL.Image)
    Args:
        pdf_path_or_bytes: PDF 파일 경로(str) 또는 bytes (io.BytesIO 가능)
        page_num: 페이지 번호 (0부터 시작)
        dpi: 원본 이미지 해상도 (낮으면 속도/용량↓)
        thumbnail_size: (width, height) 최대 크기(비율유지)
    """
    try:
        # 파일 경로 또는 바이트/버퍼 구분
        if isinstance(pdf_path_or_bytes, str):
            doc = fitz.open(pdf_path_or_bytes)
        else:
            # io.BytesIO 또는 bytes
            if isinstance(pdf_path_or_bytes, bytes):
                pdf_path_or_bytes = io.BytesIO(pdf_path_or_bytes)
            doc = fitz.open(stream=pdf_path_or_bytes, filetype="pdf")

        if not doc or page_num < 0 or page_num >= len(doc):
            return None
        
        page = doc.load_page(page_num)
        zoom = dpi / 72  # DPI 설정
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)

        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        # --- 썸네일 변환 ---
        img.thumbnail(thumbnail_size)

        doc.close()
        return img
    except Exception as e:
        logger.error(f"PDF 미리보기 생성 오류 ({pdf_path_or_bytes}, 페이지 {page_num}): {e}")
        return None


# 특정 날짜의 데이터를 S3에서 로드하는 함수
@st.cache_data(ttl=3600) # 캐시 추가: 1시간 동안 결과 유지
def load_data_for_date(date_str):
    """특정 날짜의 메타데이터, PDF 경로, OCR 결과 등을 S3에서 로드하여 세션 상태에 저장""" 
    s3_handler = S3Handler()
    data_loaded = False
    metadata = None # 메타데이터 변수 초기화
    
    # 1. 메타데이터 로드 (PDF, OCR 결과 등)
    metadata_result = s3_handler.load_metadata(date_str)
    if metadata_result["status"] == "success":
        metadata = metadata_result["data"]
        # PDF 키가 있는 경우에만 세션 상태 업데이트
        if "pdf_key" in metadata:
            st.session_state.pdf_paths_by_date[date_str] = metadata["pdf_key"]

        else:
            logger.warning(f"****** DEBUG: 메타데이터에 PDF 키 없음")
        
        # 부서-페이지 튜플 목록 로드
        if "departments_with_pages" in metadata:
            dept_page_tuples = metadata["departments_with_pages"]
            st.session_state.dept_page_tuples_by_date[date_str] = dept_page_tuples
            st.session_state.departments_with_pages_by_date[date_str] = dept_page_tuples
            logger.info(f"****** DEBUG: 날짜 {date_str}의 부서-페이지 정보 로드 성공: {len(dept_page_tuples)}개 항목")
        else:
            dept_page_tuples = [] # 없을 경우 빈 리스트로 초기화
            logger.warning(f"****** DEBUG: 메타데이터에 부서-페이지 정보 없음")
        
        # OCR 결과 로드
        ocr_text_result = s3_handler.load_ocr_text(date_str)
        if ocr_text_result["status"] == "success":
            ocr_text_list = ocr_text_result["data"]
            ocr_result = {
                "status": "success",
                "ocr_text": ocr_text_list,
                "departments_with_pages": dept_page_tuples
            }
            st.session_state.ocr_results_by_date[date_str] = ocr_result

            
            # 부서별 OCR 코드 집계 후 세션에 저장
            logger.debug(f"****** DEBUG: 부서별 OCR 코드 집계 시작 (페이지 튜플 수: {len(dept_page_tuples)})")
            try:
                codes_map = data_analyzer.aggregate_ocr_results_by_department(
                    ocr_text_list, dept_page_tuples
                )
                logger.debug(f"****** DEBUG: OCR 코드 집계 시도 결과: {codes_map.get('status', 'N/A')}")
                if codes_map.get('status') == 'success':
                    if 'aggregated_ocr_items_by_date' not in st.session_state:
                        st.session_state['aggregated_ocr_items_by_date'] = {}
                    items_by_dept = {dept: data['items'] for dept, data in codes_map.get('data', {}).items()}
                    st.session_state['aggregated_ocr_items_by_date'][date_str] = items_by_dept
                    logger.debug(f"****** DEBUG: 부서별 OCR 코드 집계 저장 성공: {len(items_by_dept)}개 부서")
                else:
                    logger.error(f"****** DEBUG: 부서별 OCR 코드 집계 실패: {codes_map.get('message', '알 수 없는 오류')}")
            except Exception as agg_e:
                logger.error(f"****** DEBUG: 부서별 OCR 코드 집계 중 예외 발생: {agg_e}", exc_info=True)
        else:
            logger.warning(f"****** DEBUG: OCR 텍스트를 찾을 수 없음")

        data_loaded = True # 메타데이터 로드 성공 시 True로 설정
        logger.debug(f"****** DEBUG: 메타데이터 기반 로드 성공")
        
        # PDF 데이터 다운로드 시도 (S3에서)
        if "pdf_key" in metadata:
            logger.debug(f"****** DEBUG: PDF 파일 다운로드 시도 (키: {metadata['pdf_key']})")
            try:
                pdf_result = s3_handler.download_file(metadata["pdf_key"])
                logger.debug(f"****** DEBUG: PDF 파일 다운로드 결과: {pdf_result['status']}")
                if pdf_result["status"] == "success":
                    # 경로가 이미 저장되었는지 다시 확인 불필요 (위에서 이미 저장됨)
                    logger.debug(f"****** DEBUG: PDF 파일 다운로드 성공")
                # 다운로드 실패 시 별도 처리 없음 (경고만 로깅됨)
            except Exception as pdf_download_e:
                logger.error(f"****** DEBUG: PDF 파일 다운로드 중 예외 발생: {pdf_download_e}", exc_info=True)
    else:
        logger.warning(f"****** DEBUG: 메타데이터 로드 실패 또는 찾을 수 없음")
    
    # 3. 엑셀 데이터 로드 (세션에 없거나 비어있는 경우)
    if 'excel_data' not in st.session_state or st.session_state.excel_data is None or st.session_state.excel_data.empty:
        logger.debug(f"****** DEBUG: 세션에 엑셀 데이터 없음. 메타데이터에서 로드 시도")
        # 메타데이터가 성공적으로 로드되었고, excel_key가 있는지 확인
        if metadata and "excel_key" in metadata:
            excel_key = metadata["excel_key"]
            logger.debug(f"****** DEBUG: 메타데이터에서 엑셀 키 '{excel_key}' 발견. 다운로드 시도")
            excel_result = s3_handler.download_file(excel_key)
            logger.debug(f"****** DEBUG: 엑셀 파일 다운로드 결과: {excel_result['status']}")
            if excel_result["status"] == "success":
                try:
                    excel_buffer_pd = io.BytesIO(excel_result["data"])
                    excel_buffer_pd.seek(0)
                    is_cumulative = "latest/cumulative_excel.xlsx" in excel_key
                    logger.debug(f"****** DEBUG: data_analyzer.load_excel_data 호출 (누적: {is_cumulative})")
                    excel_data_result = data_analyzer.load_excel_data(excel_buffer_pd, is_cumulative_flag=is_cumulative)
                    logger.debug(f"****** DEBUG: load_excel_data 결과: {excel_data_result['status']}")
                    if excel_data_result["status"] == "success":
                        st.session_state.excel_data = excel_data_result["data"]
                        st.session_state.standardized_excel_dates = sorted(
                            st.session_state.excel_data['날짜'].astype(str).unique()
                        )
                        logger.debug(f"****** DEBUG: S3에서 엑셀 데이터 로드 및 파싱 성공 ({len(st.session_state.excel_data)} 행)")
                        data_loaded = True # 엑셀 로드 성공 시 True 보장
                    else:
                        logger.error(f"****** DEBUG: 엑셀 데이터 파싱 실패: {excel_data_result.get('message', 'N/A')}")
                except Exception as excel_proc_e:
                    logger.error(f"****** DEBUG: 엑셀 데이터 처리 중 예외 발생: {excel_proc_e}", exc_info=True)
            else:
                logger.error(f"****** DEBUG: S3 엑셀 파일 다운로드 실패: {excel_result.get('message', 'N/A')}")
        else:
            logger.warning(f"****** DEBUG: 메타데이터가 없거나 'excel_key'가 없어 S3 엑셀 로드 불가")
    else:
        logger.debug(f"****** DEBUG: 세션에 이미 엑셀 데이터 존재")
        data_loaded = True # 세션에 이미 있으면 로드된 것으로 간주

    logger.debug(f"****** DEBUG: load_data_for_date 종료 (최종 data_loaded: {data_loaded})")
    # 반환 형식 변경: 불리언 대신 딕셔너리 반환
    if data_loaded:
        # 엑셀 데이터가 로드되었는지 다시 확인 후 반환
        final_excel_data = st.session_state.get('excel_data', pd.DataFrame())
        return {"status": "success", "data": final_excel_data}
    else:
        return {"status": "error", "message": f"날짜 {date_str}의 데이터를 로드할 수 없습니다."}

# PDF 파일을 표시하는 함수
def display_pdf(file_path):
    try:
        with open(file_path, "rb") as f:
            base64_pdf = base64.b64encode(f.read()).decode('utf-8')
        
        pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600" type="application/pdf"></iframe>'
        st.markdown(pdf_display, unsafe_allow_html=True)
    except Exception as e:
        st.error(f"PDF 표시 오류: {e}")

# PDF OCR 진행률 콜백
def progress_callback(current, total):
    progress_bar = st.session_state.get('progress_bar')
    if progress_bar is not None:
        progress_bar.progress(current / total)
        
        # 현재 처리 중인 페이지 정보 업데이트
        progress_text = st.session_state.get('progress_text')
        if progress_text is not None:
            progress_text.text(f"페이지 처리 중... {current}/{total}")



# --- S3 연결 확인 함수 --- 
def check_s3_connection():
    """S3 연결 상태 확인"""
    try:
        s3_client = get_s3_client()
        if s3_client is None:
            return False
        
        # 버킷 접근 테스트
        s3_client.head_bucket(Bucket=S3_BUCKET)
        logger.info(f"S3 버킷 '{S3_BUCKET}' 연결 성공")
        return True
    except Exception as e:
        logger.error(f"S3 연결 확인 실패: {e}")
        return False

# --- 물품 DB 로드 함수 (S3) --- 
@st.cache_data(ttl=3600)  # 1시간 캐시
def load_item_db_from_s3():
    """S3에서 물품 DB 파일 로드"""
    try:
        s3_client = get_s3_client()
        db_key = f"{S3_DIRS['DB']}db.xlsx"
        
        try:
            # S3에서 DB 파일 가져오기
            response = s3_client.get_object(Bucket=S3_BUCKET, Key=db_key)
            
            # 임시 파일로 저장
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_file.write(response['Body'].read())
                temp_path = temp_file.name
            
            # DB 로드
            item_db = data_analyzer.load_item_db(temp_path)
            
            # 임시 파일 삭제
            os.unlink(temp_path)
            
            logger.info("S3에서 물품 DB 파일 로드 성공")
            return item_db
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'NoSuchKey':
                logger.warning("S3에 물품 DB 파일이 없습니다.")
                return None
            raise e
            
    except Exception as e:
        logger.error(f"물품 DB 로드 실패: {e}")
        return None

# --- 물품 DB 업로드 함수 (S3) --- 
def upload_db_to_s3(file):
    """물품 DB 파일을 S3에 업로드"""
    try:
        s3_client = get_s3_client()
        db_key = f"{S3_DIRS['DB']}db.xlsx"
        
        # 파일 업로드
        s3_client.upload_fileobj(file, S3_BUCKET, db_key)
        logger.info("물품 DB 파일 S3 업로드 성공")
        return True
        
    except Exception as e:
        logger.error(f"물품 DB 파일 업로드 실패: {e}")
        return False

# pdf 처리 핵심 함수
def display_pdf_section(selected_date, sel_dept, tab_prefix="pdf_tab"):
    """
    부서별 PDF 섹션: 모든 페이지 썸네일을 한 번에 표시, 체크박스로 선택, 선택한 이미지만 S3+엑셀 저장
    """
    try:
        # S3에서 PDF 원본 다운로드
        s3_handler = S3Handler()
        pdf_key = st.session_state.pdf_paths_by_date.get(selected_date)
        if not pdf_key:
            st.warning(f"선택된 날짜({selected_date})의 PDF 파일 경로가 없습니다.")
            return

        pdf_result = s3_handler.download_file(pdf_key)
        if pdf_result["status"] != "success":
            st.error("PDF 다운로드 실패.")
            return
        pdf_bytes = pdf_result["data"]
        dept_pages = get_department_pages(selected_date, sel_dept)
        if not dept_pages:
            st.info(f"'{sel_dept}' 부서의 PDF 페이지 정보가 없습니다.")
            return

        st.subheader(f"{selected_date} {sel_dept} 미리보기 (썸네일, 다중 선택)")
        
        # Form을 사용하여 체크박스 상태 변경 시 새로고침 방지
        with st.form(key=f"{tab_prefix}_{selected_date}_image_selection_form"):
            cols = st.columns(2)
            page_checkbox_keys = []
            page_img_objs = []

            for idx, page_num in enumerate(sorted(dept_pages)):
                with cols[idx % 2]:
                    img = extract_pdf_preview(io.BytesIO(pdf_bytes), page_num-1, dpi=120, thumbnail_size=(700, 1000))
                    if img is not None:
                        st.image(img, caption=f"p.{page_num}", width=650)
                        cb_key = f"{tab_prefix}_{selected_date}_{page_num}"
                        
                        # 체크박스 표시 (Form 내부에서 새로고침 없이 동작)
                        checked = st.checkbox(
                            f"페이지 {page_num} 선택",
                            key=cb_key,
                            value=False
                        )

                        page_checkbox_keys.append((cb_key, page_num, img, checked))

            # Form 제출 버튼
            submitted = st.form_submit_button("선택한 이미지를 엑셀로 저장")
            
            if submitted:
                # --- 선택된 이미지만 추림 (Form 제출 시점의 상태 사용) ---
                selected_imgs = []
                checkbox_status = {}
                
                for cb_key, pg, img, is_checked in page_checkbox_keys:
                    checkbox_status[f"페이지 {pg}"] = is_checked
                    if is_checked:
                        selected_imgs.append((pg, img))

                # 디버깅 정보 표시
                st.write(f"**선택 결과:**")
                st.write(f"- 총 페이지 수: {len(page_checkbox_keys)}")
                st.write(f"- 선택된 이미지 수: {len(selected_imgs)}")
                st.write(f"- 체크박스 상태: {checkbox_status}")

                if not selected_imgs:
                    st.warning("저장할 이미지를 선택해주세요. 체크박스를 클릭하여 이미지를 선택한 후 버튼을 눌러주세요.")
                else:
                    saved_count = 0
                    error_count = 0
                    
                    with st.spinner(f"{len(selected_imgs)}개 이미지를 저장하는 중..."):
                        for page_num, img_obj in selected_imgs:
                            try:
                                save_result = s3_handler.save_pdf_preview_image(
                                    selected_date, sel_dept, page_num, img_obj
                                )
                                if save_result.get("status") == "success":
                                    saved_count += 1
                                    logger.info(f"이미지 저장 성공: {sel_dept} 페이지 {page_num}")
                                else:
                                    error_count += 1
                                    logger.error(f"이미지 저장 실패: {sel_dept} 페이지 {page_num} - {save_result.get('message')}")
                            except Exception as e:
                                error_count += 1
                                logger.error(f"이미지 저장 중 예외 발생: {sel_dept} 페이지 {page_num} - {e}")
                    
                    # 결과 메시지
                    if saved_count > 0:
                        st.success(f"✅ {saved_count}개 미리보기 이미지를 S3에 저장 완료! 엑셀 다운로드 시 자동 삽입됩니다.")
                        if error_count > 0:
                            st.warning(f"⚠️ {error_count}개 이미지 저장 실패")
                    else:
                        if error_count > 0:
                            st.error(f"❌ 모든 이미지 저장 실패 ({error_count}개)")
                        else:
                            st.warning("저장할 이미지가 없습니다. 체크박스를 선택했는지 확인해주세요.")

    except Exception as e:
        logger.error(f"PDF 섹션 처리 중 오류: {e}", exc_info=True)
        st.error("PDF 섹션을 처리하는 중 오류가 발생했습니다.")


# --- 헤더 행 찾기 함수 --- 
def find_header_row(df):
    """
    1행부터 10행 사이에서 부서명, 물품코드, 청구량, 수령량이 모두 포함된 행을 찾아 헤더 행으로 반환합니다.
    """
    # 확인할 키워드들 (물품코드 추가)
    keywords = ['부서명', '물품코드', '청구량', '수령량']

    # 처음 10행만 검사 (더 적은 행을 가진 경우 모든 행 검사)
    max_rows = min(10, len(df))

    for i in range(max_rows):
        try:
            row = df.iloc[i].astype(str)
            row_values = [str(val).lower().strip() for val in row.values] # 공백 제거 추가

            # 모든 키워드가 이 행에 포함되어 있는지 확인
            if all(any(keyword in val for val in row_values) for keyword in keywords):
                return i
        except IndexError:
            # 행 인덱스가 범위를 벗어나는 경우 중단
            break

    # 못 찾으면 기본값 0 반환 (또는 에러 처리)
    # logger.warning("헤더 행을 찾지 못했습니다. 기본값 0을 사용합니다.")
    return 0 # 현재 로직 유지 시 0 반환, 혹은 에러 반환 고려

# --- PDF 미리보기 엑셀 저장 함수 --- 
def save_pdf_preview_to_excel(selected_date, sel_dept, page_num, img: Image.Image, excel_path=None):
    """PDF 미리보기 이미지를 S3에 저장하고 메타데이터에 기록합니다."""
    try:
        s3_handler = S3Handler()
        
        # PIL Image 객체를 직접 전달
        result = s3_handler.save_pdf_preview_image(selected_date, sel_dept, page_num, img)

        if result["status"] == "success":
            return {"status": "success", "message": f"이미지가 저장되었습니다. 엑셀 다운로드 시 '{sel_dept}' 시트에 삽입됩니다."}
        else:
            return {"status": "error", "message": result.get("message", "알 수 없는 오류")}
    
    except Exception as e:
        logger.error(f"이미지 저장 및 메타데이터 기록 중 오류: {e}", exc_info=True)
        return {"status": "error", "message": f"이미지 저장 처리 중 오류 발생: {str(e)}"}

# --- 부서별 엑셀 다운로드 함수 (Openpyxl 단독 사용으로 수정) --- 
def download_department_excel(selected_dates):
    """
    선택한 여러 날짜의 데이터를 하나로 합쳐
    각 부서별로 시트(데이터+이미지)를 생성하여 엑셀로 반환
    """
    try:
        s3_handler = S3Handler()

        # 1. 기존 통합 mismatches_full.json 로드 (통합 작업 없이)
        df_full = s3_handler.load_full_mismatches()
        
        if df_full is None or df_full.empty:
            wb = Workbook()
            ws = wb.active
            ws.title = "데이터 없음"
            ws.cell(row=1, column=1, value="선택한 날짜에 해당하는 데이터 없음")
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue(), "부서별_통계_데이터없음.xlsx"

        # 2. 완료 처리된 항목 필터링 (세션 상태 사용)
        try:
            # 세션 상태에서 완료 로그 가져오기 (S3 로딩 없음)
            completion_logs = st.session_state.get('completion_logs', [])
            filtered_df = filter_completed_items(df_full, completion_logs) if completion_logs else df_full
        except Exception as e:
            logger.warning(f"완료처리 기록 필터링 오류: {e}")
            filtered_df = df_full

        # 3. 선택한 날짜로 필터링
        all_excels = []
        for dt in selected_dates:
            # 날짜 컬럼 변환 및 필터링
            if not pd.api.types.is_datetime64_any_dtype(filtered_df['날짜']):
                filtered_df['날짜'] = pd.to_datetime(filtered_df['날짜'], errors='coerce')
            
            df = filtered_df[filtered_df['날짜'].dt.strftime('%Y-%m-%d') == dt].copy()
            if not df.empty:
                all_excels.append(df)
                
        if not all_excels:
            wb = Workbook()
            ws = wb.active
            ws.title = "데이터 없음"
            ws.cell(row=1, column=1, value="선택한 날짜에 해당하는 데이터 없음")
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue(), "부서별_통계_데이터없음.xlsx"

        excel_df = pd.concat(all_excels, ignore_index=True)

        # 4. 선택 날짜의 모든 이미지 취합 (메타데이터 기준)
        dept_images = {}
        missing_depts_with_images = set()  # 누락된 부서 추적
        
        for dt in selected_dates:
            metadata_result = s3_handler.load_metadata(dt)
            metadata = metadata_result.get("data", {}) if metadata_result.get("status") == "success" else {}
            preview_images = metadata.get("preview_images", [])
            
            # 해당 날짜의 엑셀 부서 목록 가져오기
            excel_depts_for_date = set()
            if not excel_df.empty:
                date_filtered_excel = excel_df[excel_df['날짜'].dt.strftime('%Y-%m-%d') == dt]
                if not date_filtered_excel.empty and '부서명' in date_filtered_excel.columns:
                    excel_depts_for_date = set(date_filtered_excel['부서명'].unique())
            
            # PDF 부서 목록 가져오기 (departments_with_pages_by_date에서)
            pdf_depts_for_date = set()
            if dt in st.session_state.get('departments_with_pages_by_date', {}):
                dept_page_tuples = st.session_state.departments_with_pages_by_date[dt]
                pdf_depts_for_date = {dept for dept, page in dept_page_tuples if dept}
            
            # 누락된 부서 찾기 (PDF에만 있고 엑셀에 없는 부서)
            missing_depts_for_date = pdf_depts_for_date - excel_depts_for_date
            missing_depts_with_images.update(missing_depts_for_date)
            
            # 모든 이미지 정보 수집
            for img_info in preview_images:
                dept = img_info.get("dept")
                if dept: 
                    # 날짜 정보 추가
                    img_info_with_date = img_info.copy()
                    img_info_with_date['date'] = dt
                    dept_images.setdefault(dept, []).append(img_info_with_date)

        # 5. 부서별로 시트 생성 (엑셀 부서 + 이미지가 있는 모든 부서 포함)
        # 엑셀에 있는 부서 목록
        excel_depts = set()
        if not excel_df.empty and '부서명' in excel_df.columns:
            excel_depts = set(excel_df['부서명'].unique())
        
        # 이미지가 있는 모든 부서 (PDF에서 추출된 부서 포함)
        image_depts = set(dept_images.keys())
        
        # 전체 부서 목록 (엑셀 부서 + 이미지 부서)
        all_depts = excel_depts | image_depts
        
        # 누락된 부서 확인 및 로깅
        missing_depts_final = missing_depts_with_images & image_depts
        existing_depts = excel_depts & image_depts
        
        logger.info(f"엑셀 시트 생성 - 총 부서 수: {len(all_depts)}")
        logger.info(f"  - 엑셀에 있는 부서: {len(excel_depts)}개")
        logger.info(f"  - 이미지가 있는 부서: {len(image_depts)}개")
        logger.info(f"  - 누락된 부서 (새 시트 생성): {len(missing_depts_final)}개")
        if missing_depts_final:
            logger.info(f"  - 누락된 부서 목록: {', '.join(sorted(missing_depts_final))}")
        
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])  # 기본 시트 제거

        headers = ['날짜', '부서명', '물품코드', '물품명', '청구량', '수령량', '차이', '누락']

        for dept in sorted(list(all_depts)):
            # 안전한 시트명 생성 (엑셀 시트명 제한사항 고려)
            safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', str(dept))[:31]
            
            # 시트명 중복 방지 (같은 이름의 시트가 이미 있는지 확인)
            original_name = safe_sheet_name
            counter = 1
            while safe_sheet_name in [ws.title for ws in wb.worksheets]:
                safe_sheet_name = f"{original_name[:28]}_{counter}"
                counter += 1
            
            # 새 시트 생성
            ws = wb.create_sheet(safe_sheet_name)
            
            # 누락된 부서인지 확인
            is_missing_dept = dept in missing_depts_with_images
            has_excel_data = dept in excel_depts
            
            # 시트 생성 로깅
            if is_missing_dept:
                logger.info(f"누락된 부서 '{dept}' 시트 생성: {safe_sheet_name}")
            else:
                logger.debug(f"일반 부서 '{dept}' 시트 생성: {safe_sheet_name}")
            
            # 데이터: 선택한 모든 날짜의 해당 부서 데이터만 추출
            dept_df_export = pd.DataFrame(columns=headers)
            if has_excel_data and not excel_df.empty:
                dept_df_filtered = excel_df[excel_df['부서명'] == dept].copy()
                for col in headers:
                    if col not in dept_df_filtered.columns:
                        if col == '차이':
                            dept_df_filtered[col] = dept_df_filtered.get('수령량', 0) - dept_df_filtered.get('청구량', 0)
                        else:
                            dept_df_filtered[col] = ''
                mask = (dept_df_filtered['차이'] == 1) & (dept_df_filtered['청구량'] == 0) & (dept_df_filtered['수령량'] == 1)
                dept_df_filtered.loc[mask, '누락'] = '누락'
                dept_df_filtered['누락'] = dept_df_filtered['누락'].fillna('')
                # 날짜 칼럼 포맷
                if '날짜' in dept_df_filtered.columns:
                    dept_df_filtered['날짜'] = pd.to_datetime(dept_df_filtered['날짜'], errors='coerce').dt.strftime('%Y-%m-%d')
                dept_df_export = dept_df_filtered[headers]
                logger.debug(f"부서 '{dept}' 엑셀 데이터: {len(dept_df_export)}행")
            elif is_missing_dept:
                logger.info(f"누락된 부서 '{dept}': 엑셀 데이터 없음, 이미지만 포함")
            
            # 헤더 작성
            ws.append(headers)
            
            # 데이터 작성
            if not dept_df_export.empty:
                for r in dataframe_to_rows(dept_df_export, index=False, header=False):
                    ws.append(r)
            else:
                # 누락된 부서의 경우 특별한 메시지 표시
                if is_missing_dept:
                    ws.cell(row=2, column=1, value="⚠️ 누락된 부서")
                    ws.cell(row=2, column=2, value=dept)
                    ws.cell(row=3, column=1, value="상태")
                    ws.cell(row=3, column=2, value="PDF 인수증에만 있고 엑셀 데이터에는 없는 부서")
                    ws.cell(row=4, column=1, value="발견 날짜")
                    # 해당 부서가 발견된 날짜들 표시
                    dept_dates = []
                    for img_info in dept_images.get(dept, []):
                        date_info = img_info.get('date', '')
                        if date_info and date_info not in dept_dates:
                            dept_dates.append(date_info)
                    ws.cell(row=4, column=2, value=", ".join(sorted(dept_dates)) if dept_dates else "알 수 없음")
                    ws.cell(row=5, column=1, value="조치 필요")
                    ws.cell(row=5, column=2, value="아래 인수증 이미지를 확인하여 누락된 데이터를 엑셀에 추가하세요.")
                    ws.cell(row=6, column=1, value="이미지 수")
                    ws.cell(row=6, column=2, value=f"{len(dept_images.get(dept, []))}개")
                else:
                    ws.cell(row=2, column=1, value=f"데이터 없음")
                    ws.cell(row=2, column=2, value="이 부서는 선택한 기간에 데이터가 없습니다.")

            # 이미지 삽입
            images = dept_images.get(dept, [])
            if images:
                last_data_col = len(headers)
                image_col_start = last_data_col + 2
                max_images_per_row = 2
                
                # 누락된 부서의 경우 특별한 헤더 표시
                if is_missing_dept:
                    ws.cell(row=1, column=image_col_start, value="⚠️ 누락 부서 인수증 이미지")
                    # 누락된 부서 설명 추가
                    ws.cell(row=2, column=image_col_start, value="이 부서는 PDF에만 있고")
                    ws.cell(row=3, column=image_col_start, value="엑셀 데이터에는 없습니다.")
                    ws.cell(row=4, column=image_col_start, value="아래 이미지를 확인하여")
                    ws.cell(row=5, column=image_col_start, value="누락된 데이터를 추가하세요.")
                    ws.cell(row=6, column=image_col_start, value="자동 생성된 시트입니다.")
                    image_start_row = 8  # 더 많은 정보가 추가되었으므로 시작 행 조정
                else:
                    ws.cell(row=1, column=image_col_start, value="인수증 이미지")
                    image_start_row = 2
                
                for i, img_info in enumerate(images):
                    try:
                        img_bytes = get_pdf_preview_image_from_s3(img_info["file_key"])
                        if not img_bytes:
                            continue  # 이미지가 없으면 건너뜀
                        xl_img = XLImage(io.BytesIO(img_bytes))
                        xl_img.width = 350
                        xl_img.height = 500
                        row_idx = i // max_images_per_row
                        col_idx = i % max_images_per_row
                        col_pos = image_col_start + (col_idx * 4)
                        current_row_for_images = image_start_row + (row_idx * 15)
                        
                        # 이미지 정보 (누락된 부서의 경우 더 자세한 정보 표시)
                        date_info = img_info.get('date', '')
                        page_info = img_info.get('page', '?')
                        if is_missing_dept:
                            ws.cell(row=current_row_for_images, column=col_pos, 
                                   value=f"⚠️ 누락부서: {dept}")
                            ws.cell(row=current_row_for_images+1, column=col_pos, 
                                   value=f"날짜: {date_info}, 페이지: {page_info}")
                            img_row = current_row_for_images + 2
                        else:
                            ws.cell(row=current_row_for_images, column=col_pos, 
                                   value=f"날짜:{date_info}, 페이지:{page_info}")
                            img_row = current_row_for_images + 1
                        
                        ws.add_image(xl_img, f"{ws.cell(row=img_row, column=col_pos).coordinate}")
                    except Exception as e:
                        logger.error(f"이미지 처리/삽입 중 오류 ({dept}, {img_info.get('page')}): {e}")
                        continue

        # 열 너비
        for ws in wb.worksheets:
            std_widths = {
                'A': 12, 'B': 20, 'C': 12, 'D': 30, 'E': 10, 'F': 10, 'G': 10, 'H': 10,
            }
            for col, width in std_widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width

        # 최종 엑셀 파일 버퍼에 저장
        excel_buffer_final = io.BytesIO()
        wb.save(excel_buffer_final)
        excel_buffer_final.seek(0)
        
        # 파일명 생성 (누락된 부서 정보 포함)
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        if missing_depts_with_images:
            missing_count = len(missing_depts_with_images)
            file_name = f"부서별_통계_누락부서{missing_count}개포함_{current_time}.xlsx"
            logger.info(f"✅ 엑셀 다운로드 완료: 총 {len(all_depts)}개 부서 시트 생성")
            logger.info(f"   - 일반 부서: {len(excel_depts)}개")
            logger.info(f"   - 누락 부서 (자동 생성): {missing_count}개")
            logger.info(f"   - 누락된 부서 목록: {', '.join(sorted(missing_depts_with_images))}")
            logger.info(f"   - 파일명: {file_name}")
        else:
            file_name = f"부서별_통계_{current_time}.xlsx"
            logger.info(f"✅ 엑셀 다운로드 완료: 총 {len(all_depts)}개 부서 시트 생성 (누락 부서 없음)")
            logger.info(f"   - 파일명: {file_name}")
        
        return excel_buffer_final.getvalue(), file_name

    except Exception as e:
        logger.error(f"엑셀 다운로드(download_department_excel) 중 오류: {e}", exc_info=True)
        st.error(f"엑셀 다운로드 중 오류가 발생했습니다: {str(e)}")
        return None, None


# --- 날짜 옵션 가져오기 함수 --- 
def get_date_options():
    """처리된 날짜 목록을 반환합니다."""
    s3_handler = S3Handler()
    result = s3_handler.list_processed_dates()
    
    # 결과가 딕셔너리이고 'status'가 'success'인 경우 'dates' 키에서 날짜 목록을 가져옴
    if isinstance(result, dict):
        if result.get('status') == 'success' and 'dates' in result:
            return sorted(result['dates'], reverse=True)
        elif result.get('status') == 'error':
            logger.error(f"날짜 목록 조회 실패: {result.get('message')}")
            return []
    
    # 결과가 리스트인 경우 그대로 반환
    if isinstance(result, list):
        return sorted(result, reverse=True)
    
    # 그 외의 경우 빈 리스트 반환
    logger.warning(f"예상치 못한 날짜 목록 형식: {type(result)}")
    return []


# --- 앱 설정, 스타일, 세션 상태 초기화 등 --- 
# ... (기존 앱 설정 및 세션 상태 초기화 코드) ...

# 메인 함수
def main():
    # S3 연결 확인
    if not check_s3_connection():
        st.error("S3 연결에 실패했습니다. 관리자에게 문의하세요.")
        return
    
    # 세션 상태 초기화
    if 'pdf_paths_by_date' not in st.session_state:
        st.session_state.pdf_paths_by_date = {}
    if 'dept_page_tuples_by_date' not in st.session_state:
        st.session_state.dept_page_tuples_by_date = {}
    if 'ocr_results_by_date' not in st.session_state:
        st.session_state.ocr_results_by_date = {}
    if 'aggregated_ocr_items_by_date' not in st.session_state:
        st.session_state.aggregated_ocr_items_by_date = {}
    if 'departments_with_pages_by_date' not in st.session_state:
        st.session_state.departments_with_pages_by_date = {}
    if 'loaded_dates' not in st.session_state:
        st.session_state.loaded_dates = set()  # 이미 로드된 날짜를 추적하기 위한 세트
    if 'item_db' not in st.session_state:
        st.session_state.item_db = {}  # 물품 DB 초기화 추가

    
    # 완료 처리 로그 로드 (앱 시작 시)
    if 'completion_logs' not in st.session_state:
        try:
            s3_handler = S3Handler()
            completion_logs_result = s3_handler.load_completion_logs()
            
            if completion_logs_result["status"] == "success":
                st.session_state.completion_logs = completion_logs_result["data"]
                logger.info(f"앱 시작 시 완료 처리 로그 {len(st.session_state.completion_logs)}개 로드 성공.")
            elif completion_logs_result["status"] == "not_found":
                st.session_state.completion_logs = []
                logger.info("앱 시작 시 완료 처리 로그 파일이 존재하지 않아 빈 리스트로 초기화합니다.")
            else: # "error" 또는 기타 상태
                st.session_state.completion_logs = []
                logger.error(f"앱 시작 시 완료 처리 로그 로드 실패: {completion_logs_result.get('message', '알 수 없는 오류')}. 빈 리스트로 초기화합니다.")
        except Exception as e:
            st.session_state.completion_logs = []
            logger.error(f"앱 시작 시 완료 처리 로그 로드 중 심각한 예외 발생: {e}", exc_info=True)
    else:
        # 세션에 이미 있어도 S3에서 최신 데이터 강제 로드
        try:
            s3_handler = S3Handler()
            completion_logs_result = s3_handler.load_completion_logs()
            
            if completion_logs_result["status"] == "success":
                # S3 데이터와 세션 데이터 비교
                s3_logs = completion_logs_result["data"]
                session_logs = st.session_state.completion_logs
                
                if len(s3_logs) != len(session_logs):
                    logger.info(f"S3와 세션의 완료 로그 수가 다름 (S3: {len(s3_logs)}, 세션: {len(session_logs)}). S3 데이터로 업데이트.")
                    st.session_state.completion_logs = s3_logs
                else:
                    logger.info(f"완료 처리 로그 이미 로드됨: {len(session_logs)}개 항목")
            else:
                logger.warning(f"S3에서 완료 로그 재로드 실패: {completion_logs_result.get('message')}")
        except Exception as e:
            logger.error(f"완료 로그 재로드 중 오류: {e}")
    
    # 한글 폰트 설정
    set_korean_font()
    
    st.title("상계백병원 인수증 & 엑셀 데이터 비교 시스템")
    
    s3_handler = S3Handler()

    # --- 앱 시작 시 데이터 로드 최적화 (통합 작업 제거) ---
    if 'mismatch_data' not in st.session_state or st.session_state.mismatch_data.empty:
        logger.info("세션에 불일치 데이터가 없거나 비어있어 S3에서 로드를 시도합니다.")
        try:
            # 기존 통합 파일만 로드 (통합 작업은 하지 않음)
            full_mismatches = s3_handler.load_full_mismatches()
            if not full_mismatches.empty:
                st.session_state.mismatch_data = full_mismatches
                logger.info(f"기존 통합 mismatches_full.json 로드 완료: {len(full_mismatches)}개 항목")
            else:
                # 통합 파일이 없어도 앱 시작 시에는 통합 작업하지 않음
                st.session_state.mismatch_data = pd.DataFrame()
                logger.info("기존 통합 파일이 없어 빈 DataFrame으로 초기화")
        except Exception as e:
            logger.error(f"불일치 데이터 초기화 중 오류: {e}")
            st.session_state.mismatch_data = pd.DataFrame()
    else:
        logger.info("세션에 이미 불일치 데이터가 존재합니다. S3 로드를 건너뜁니다.")

    if 'excel_data' not in st.session_state or st.session_state.excel_data.empty:
        logger.info("세션에 엑셀 데이터가 없어 강제 리로드를 시도합니다.")
        force_reload_excel_data(s3_handler) # 앱 시작 시 항상 최신 엑셀 데이터 로드

    # 불일치 데이터 재계산은 필요한 경우에만 수행 (예: 파일 업로드 후)
    # process_files 함수 내부에서 재계산 로직 호출
    
    # 이미 처리된 날짜를 추적하는 세션 변수 추가
    if 'metadata_updated_dates' not in st.session_state:
        st.session_state.metadata_updated_dates = set()

    # PDF 키 누락된 메타데이터 수정
    if 'available_dates' in st.session_state:
        for date in st.session_state.available_dates:
            update_metadata_with_pdf(s3_handler, date)
            
    # 강제 리로드 플래그 처리
    if st.session_state.get('force_reload_mismatch', False):
        logger.info("force_reload_mismatch 플래그 감지. 불일치 데이터 재계산 및 UI 새로고침")
        recalculate_mismatches(s3_handler) # 재계산 수행
        del st.session_state.force_reload_mismatch
        st.rerun() # UI 즉시 업데이트

    # S3 저장소에서 처리된 날짜 목록 로드
    processed_dates_result = s3_handler.list_processed_dates()
    if processed_dates_result["status"] == "success" and processed_dates_result["dates"]:
        st.session_state.dates = sorted(processed_dates_result["dates"])  # 전체 날짜를 저장

        # --- 작업 기간 선택 UI 추가 ---
        st.sidebar.header("작업 기간 선택")
        col1, col2 = st.sidebar.columns(2)
        
        # 기본값 설정
        default_start = datetime.today() - timedelta(days=30)
        default_end = datetime.today()
        
        # 세션 상태 초기화 (위젯 생성 전에)
        if 'work_start_date' not in st.session_state:
            st.session_state.work_start_date = default_start.date()
        if 'work_end_date' not in st.session_state:
            st.session_state.work_end_date = default_end.date()
        
        start_date = col1.date_input(
            "시작일", 
            value=st.session_state.work_start_date, 
            key="date_input_start"
        )
        end_date = col2.date_input(
            "종료일", 
            value=st.session_state.work_end_date, 
            key="date_input_end"
        )

        # 세션 상태 업데이트
        st.session_state.work_start_date = start_date
        st.session_state.work_end_date = end_date

        if st.session_state.work_start_date > st.session_state.work_end_date:
            st.sidebar.error("시작일은 종료일보다 앞서야 합니다.")
            st.stop()

        st.session_state.available_dates = [
            d for d in st.session_state.dates
            if st.session_state.work_start_date <= datetime.strptime(d, "%Y-%m-%d").date() <= st.session_state.work_end_date
        ]

        if not st.session_state.available_dates:
            st.sidebar.warning("선택한 기간에 해당하는 날짜가 없습니다.")
            st.stop()


    # 사이드바
    with st.sidebar:
        st.header("파일 업로드")
        
        # --- 다중 엑셀 업로드 허용 ---
        excel_files = st.file_uploader(
            "엑셀 파일 업로드 (여러 개 선택 가능)", 
            type=["xlsx", "xls"],
            accept_multiple_files=True # 다중 파일 허용
        )
        # -------------------------
        
        # --- 다중 PDF 업로드 허용 ---
        pdf_files = st.file_uploader(
            "PDF 파일 업로드 (여러 개 선택 가능)",
            type=["pdf"],
            accept_multiple_files=True # 다중 파일 허용
        )
        # -------------------------
        
        # 처리 버튼
        if excel_files or pdf_files: # 둘 중 하나라도 있으면 처리 가능
            if st.button("처리 시작", key="process_button"):
                process_files(excel_files, pdf_files) # excel_files 리스트 전달
        

        
        # 리셋 버튼
        if st.button("모든 데이터 초기화", key="reset_button"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.success("모든 데이터가 초기화되었습니다.")
            
        st.markdown("---")
        st.markdown("""
        ### 사용 방법
        1. 엑셀 파일과 PDF 파일(들)을 업로드하세요.
        2. '처리 시작' 버튼을 클릭하세요.
        3. **작업 기간**을 설정하세요.
        4. 각 탭에서 세부 날짜를 선택하여 결과를 확인하세요.
           (PDF 관련 정보는 해당 날짜의 PDF가 있을 때만 표시됩니다.)
        """)

        # 물품 DB 로드 (S3 우선, 없으면 업로드 허용)
        if not st.session_state.item_db:
            # S3에서 DB 파일 로드 시도
            s3_db = load_item_db_from_s3()
            if s3_db:
                st.session_state.item_db = s3_db
                st.info("S3에서 물품 DB 파일을 로드했습니다.")
            else:
                # S3에 없으면 업로드 허용
                db_file = st.file_uploader("물품 DB 파일 업로드", type=["xlsx"], key="db_file")
                if db_file:
                    # 업로드된 파일을 S3에 저장
                    if upload_db_to_s3(db_file):
                        # 파일 포인터 위치 리셋
                        db_file.seek(0)
                        # 임시 파일로 저장하고 로드
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                            temp_file.write(db_file.read())
                            temp_path = temp_file.name
                        st.session_state.item_db = data_analyzer.load_item_db(temp_path)
                        os.unlink(temp_path)  # 임시 파일 삭제
                        st.success("물품 DB 파일이 업로드되고 로드되었습니다.")

    # 탭 생성
    tabs = st.tabs(["날짜별 작업", "부서별 통계", "완료 항목 관리"])
    
    # 선택된 날짜 확인
    if 'selected_date' not in st.session_state or not st.session_state.selected_date:
        # 사이드바에서 날짜를 선택하지 않았거나, available_dates가 없을 수 있음
        # 이 경우, 날짜별 작업 탭에서 날짜 선택을 유도하거나, 기간 내 첫 날짜를 기본으로 할 수 있음
        # 여기서는 날짜별 작업 탭에서 처리하도록 하고, main에서는 blocking하지 않음
        pass # display_mismatch_tab에서 처리

    selected_date = st.session_state.selected_date
    
    # 탭 1: 날짜별 작업
    with tabs[0]:
        display_mismatch_tab() # 인자 없이 호출

    # 탭 2: 부서별 통계
    with tabs[1]:
        display_filter_tab()
           
    # 탭 3: 완료 항목 관리
    with tabs[2]:
        display_completed_items_tab() # 새로 추가할 함수 호출


# 파일 처리 함수 (다중 PDF 처리)

def process_files(excel_files, pdf_files):
    try:
        s3_handler = S3Handler()
        processed_dates = set() # 날짜 중복 방지를 위해 set 사용
        current_excel_data = pd.DataFrame()
        cumulative_excel_key = f"{S3_DIRS['EXCEL']}latest/cumulative_excel.xlsx"
        # --- 1. 기존 누적 엑셀 데이터 로드 시도 --- 
        st.write("기존 누적 엑셀 데이터 로드를 시도합니다...")
        try:
            excel_download_result = s3_handler.download_file(cumulative_excel_key)
            if excel_download_result["status"] == "success":
                excel_buffer = io.BytesIO(excel_download_result["data"])
                # 누적 파일이므로 is_cumulative_flag=True 전달
                load_result = data_analyzer.load_excel_data(excel_buffer, is_cumulative_flag=True)
                if load_result["status"] == "success":
                    current_excel_data = load_result["data"]
                    logger.info(f"S3에서 기존 누적 엑셀 데이터 로드 성공: {len(current_excel_data)}개 행")
                    # 기존 데이터의 날짜도 processed_dates에 추가
                    if '날짜' in current_excel_data.columns:
                        processed_dates.update(current_excel_data['날짜'].astype(str).unique())
                else:
                    logger.warning(f"S3에서 다운로드한 누적 엑셀 파일 로드 실패: {load_result['message']}")
            elif excel_download_result["status"] == "not_found":
                logger.info("S3에 기존 누적 엑셀 파일이 없습니다. 새로 시작합니다.")
            else:
                logger.error(f"S3에서 누적 엑셀 파일 다운로드 실패: {excel_download_result['message']}")
        except Exception as e:
            logger.error(f"기존 누적 엑셀 데이터 로드 중 오류: {e}", exc_info=True)
            st.warning("기존 누적 엑셀 데이터를 로드하는 중 오류가 발생했습니다.")
        # -------------------------------------
        
        newly_processed_excel_files = [] # 새로 처리된 엑셀 파일명 저장
        
        # --- 2. 새로 업로드된 엑셀 파일 처리 --- 
        if excel_files:
    
            progress_bar_excel = st.progress(0)
            status_text_excel = st.empty()
            
            for i, uploaded_excel_file in enumerate(excel_files, 1):
                status_text_excel.text(f"엑셀 파일 처리 중 ({i}/{len(excel_files)}): {uploaded_excel_file.name}")
                try:
                    # 파일 읽기
                    uploaded_excel_file.seek(0)
                    excel_buffer_new = io.BytesIO(uploaded_excel_file.read())
                    uploaded_excel_file.seek(0) # 다음 사용 위해 포인터 리셋
                    
                    # 데이터 로드 (일반 파일이므로 is_cumulative_flag=False 명시)
                    logger.info(f"'{uploaded_excel_file.name}' 로드 시도 (is_cumulative=False)")
                    new_data_result = data_analyzer.load_excel_data(excel_buffer_new, is_cumulative_flag=False)
                    
                    if new_data_result["status"] == "success":
                        new_data_df = new_data_result["data"]
                        logger.info(f"엑셀 파일 '{uploaded_excel_file.name}' 로드 성공: {len(new_data_df)}개 행")
                        
                        # 기존 데이터와 병합
                        current_excel_data = pd.concat([current_excel_data, new_data_df], ignore_index=True)
                        logger.info(f"'{uploaded_excel_file.name}' 데이터 병합 후 총 {len(current_excel_data)}개 행")
                        
                        # 새로 처리된 날짜 추가
                        if '날짜' in new_data_df.columns:
                            processed_dates.update(new_data_df['날짜'].astype(str).unique())
                        
                        newly_processed_excel_files.append(uploaded_excel_file.name)
                    else:
                        st.warning(f"엑셀 파일 '{uploaded_excel_file.name}' 로드 실패: {new_data_result['message']}")
                        logger.warning(f"엑셀 파일 '{uploaded_excel_file.name}' 로드 실패, 병합 건너뜀: {new_data_result['message']}")
                        
                    newly_processed_excel_files.append(uploaded_excel_file.name)
                except Exception as e:
                    logger.error(f"엑셀 파일 '{uploaded_excel_file.name}' 처리 중 오류: {e}", exc_info=True)
                    st.error(f"엑셀 파일 '{uploaded_excel_file.name}' 처리 중 오류가 발생했습니다.")
                progress_bar_excel.progress(i / len(excel_files))
            
            status_text_excel.text("엑셀 파일 처리 완료. 중복 제거 중...")
            
            # --- 3. 중복 제거 --- 
            key_columns = ['날짜', '부서명', '물품코드']
            if all(col in current_excel_data.columns for col in key_columns):
                initial_rows = len(current_excel_data)
                current_excel_data = current_excel_data.drop_duplicates(subset=key_columns, keep='last').reset_index(drop=True)
                removed_rows = initial_rows - len(current_excel_data)
                logger.info(f"중복 데이터 제거 완료. {removed_rows}개 행 제거됨. 최종 {len(current_excel_data)}개 행.")
            else:
                logger.warning(f"중복 제거 위한 키 컬럼 부족: {key_columns}. 중복 제거 건너뜀.")
            
            # --- 4. 누적 엑셀 데이터 S3 저장 --- 
            if not current_excel_data.empty:
                try:
                    excel_output_buffer = io.BytesIO()
                    current_excel_data.to_excel(excel_output_buffer, index=False)
                    excel_output_buffer.seek(0)
                    
                    # 해시 계산 (선택적, 메타데이터용)
                    cumulative_excel_hash_result = s3_handler.get_file_hash(excel_output_buffer)
                    cumulative_excel_hash = cumulative_excel_hash_result.get("hash") if cumulative_excel_hash_result["status"] == "success" else None
                    
                    excel_output_buffer.seek(0)
                    upload_result = s3_handler.upload_file(
                        excel_output_buffer, 
                        "latest", # 날짜 대신 'latest' 사용
                        "cumulative_excel.xlsx", # 고정 파일명 사용
                        'EXCEL' # 디렉토리 타입
                    )
                    if upload_result["status"] == "success":
                        cumulative_excel_key = upload_result["key"] # 실제 저장된 키 업데이트
                        logger.info(f"누적 엑셀 데이터를 S3에 저장했습니다: {cumulative_excel_key}")
                    else:
                        st.error(f"누적 엑셀 데이터 S3 저장 실패: {upload_result['message']}")
                except Exception as e:
                    logger.error(f"누적 엑셀 데이터 S3 저장 중 오류: {e}", exc_info=True)
                    st.error("누적 엑셀 데이터를 S3에 저장하는 중 오류가 발생했습니다.")
            else:
                logger.warning("저장할 누적 엑셀 데이터가 없습니다.")
            
            # --- 5. 세션 상태 업데이트 --- 
            st.session_state.excel_data = current_excel_data
            if not current_excel_data.empty:
                st.session_state.standardized_excel_dates = sorted(
                    current_excel_data['날짜'].astype(str).unique()
                )
                # logger.info(f"세션 엑셀 날짜 업데이트: {len(st.session_state.standardized_excel_dates)}개") # 중복 로그 제거
            else:
                st.session_state.standardized_excel_dates = []
                
            st.success(f"{len(excel_files)}개 엑셀 파일 처리가 완료되었습니다.")
            progress_bar_excel.empty()
            status_text_excel.empty()
        # -------------------------------------

        # --- PDF 파일 처리 (기존 로직과 유사하게 진행, processed_dates 사용) ---
        # ... (기존 PDF 처리 로직, 단 processed_dates.update(...) 사용) ...
        if pdf_files:
            st.markdown("---")
            st.subheader("PDF 파일 처리")
            total_pdfs = len(pdf_files)

            
            progress_bar_pdf = st.progress(0)
            status_text_pdf = st.empty()

            for i, pdf_file in enumerate(pdf_files, 1):
                status_text_pdf.write(f"PDF 파일 처리 중 ({i}/{total_pdfs}): {pdf_file.name}")

                # 1. 파일 내용 해시 계산
                pdf_hash_result = s3_handler.get_file_hash(pdf_file)
                if pdf_hash_result["status"] != "success":
                    st.error(f"PDF 파일 해시 계산 실패: {pdf_hash_result['message']}")
                    continue
                pdf_hash = pdf_hash_result["hash"]

                # 2. PDF 파일명에서 날짜 추출
                pdf_filename = pdf_file.name
                extracted_date = standardize_date(pdf_filename)
                
                if extracted_date == pdf_filename:
                    st.warning(f"'{pdf_filename}' 파일명에서 날짜를 추출할 수 없어 현재 날짜를 사용합니다.")
                    pdf_date = datetime.now().strftime('%Y-%m-%d')
                else:
                    pdf_date = extracted_date
                    logger.info(f"PDF 파일명 '{pdf_filename}'에서 날짜 추출: {pdf_date}")

                # 3. 해시값으로 이미 처리된 PDF인지 확인
                exists_result = s3_handler.check_file_exists(pdf_date, pdf_hash, "PDF")
                if exists_result["status"] == "success" and exists_result["exists"]:
                    metadata = exists_result["metadata"]
                    st.info(f"'{pdf_file.name}' ({pdf_date}) 파일은 이미 처리되어 있습니다. 기존 결과를 불러옵니다.")
                    
                    ocr_text_result = s3_handler.load_ocr_text(pdf_date)
                    if ocr_text_result["status"] == "success":
                        ocr_result = {
                            "status": "success",
                            "ocr_text": ocr_text_result["data"],
                            "departments_with_pages": metadata.get("departments_with_pages", [])
                        }
                        st.session_state.pdf_paths_by_date[pdf_date] = metadata["pdf_key"]
                        st.session_state.ocr_results_by_date[pdf_date] = ocr_result
                        st.session_state.dept_page_tuples_by_date[pdf_date] = metadata.get("departments_with_pages", [])
                        
                        # departments_with_pages_by_date 세션 상태 명시적 업데이트 추가
                        dept_pages = metadata.get("departments_with_pages", [])
                        st.session_state.departments_with_pages_by_date[pdf_date] = dept_pages
                        # logger.info(f"기존 PDF 처리 결과 로드 - 날짜 {pdf_date}의 부서-페이지 정보: {len(dept_pages)}개 항목") # 중복 로그 제거
                        
                        processed_dates.add(pdf_date) # 처리된 날짜 set에 추가
                        # logger.info(f"기존 PDF 처리 결과 로드 완료: {pdf_date}") # 중복 로그 제거
                    else:
                        st.warning(f"OCR 텍스트를 로드할 수 없습니다. 파일을 다시 처리합니다.")
                        exists_result["exists"] = False
                
                if not exists_result.get("exists", False):
                    pdf_file.seek(0)
                    pdf_bytes = pdf_file.read()
                    pdf_buffer_s3 = io.BytesIO(pdf_bytes)
                    pdf_buffer_s3.seek(0)
                    pdf_upload_result = s3_handler.upload_file(
                        pdf_buffer_s3,
                        pdf_date,
                        pdf_file.name,
                        'PDF'
                    )
                    if pdf_upload_result["status"] != "success":
                        st.error(f"PDF 파일 업로드 실패: {pdf_upload_result['message']}")
                        continue
                    
                    pdf_buffer_proc = io.BytesIO(pdf_bytes)
                    pdf_buffer_proc.seek(0)

                    ocr_result = pdf3_module.process_pdf(pdf_buffer_proc)
                    
                    if ocr_result["status"] == "success":
                        ocr_text_save_result = s3_handler.save_ocr_text(pdf_date, ocr_result["ocr_text"])
                        if ocr_text_save_result["status"] != "success":
                            st.warning(f"OCR 텍스트 저장 실패: {ocr_text_save_result['message']}")
                        
                        departments_with_pages = ocr_result.get("departments_with_pages", [])
                        metadata = {
                            "pdf_key": pdf_upload_result["key"],
                            "pdf_hash": pdf_hash,
                            "pdf_filename": pdf_file.name,
                            "ocr_pages": len(ocr_result["ocr_text"]),
                            "departments_with_pages": departments_with_pages,
                            "processed_date": datetime.now().isoformat()
                            # 엑셀 관련 정보는 아래 메타데이터 업데이트에서 추가
                        }
                        # 메타데이터 저장 (임시, 아래에서 덮어쓸 수 있음)
                        s3_handler.save_metadata(pdf_date, metadata) 

                        st.session_state.pdf_paths_by_date[pdf_date] = pdf_upload_result["key"]
                        st.session_state.ocr_results_by_date[pdf_date] = ocr_result
                        processed_dates.add(pdf_date) # 처리된 날짜 set에 추가
                        st.success(f"'{pdf_file.name}' 파일 처리가 완료되었습니다.")
                    else:
                        st.error(f"'{pdf_file.name}' OCR 처리 실패: {ocr_result.get('message', '알 수 없는 오류')}")
                
                        # departments_with_pages_by_date 세션 상태 명시적 업데이트 추가
                        if "departments_with_pages" in metadata:
                            st.session_state.departments_with_pages_by_date[pdf_date] = metadata["departments_with_pages"]
                
                progress_bar_pdf.progress(i / total_pdfs)

            status_text_pdf.empty()
            progress_bar_pdf.empty()
            st.success(f"총 {total_pdfs}개의 PDF 파일 처리가 완료되었습니다.")
        # -------------------------------------
        
        # --- 6. 누적 불일치 데이터 계산 및 저장 ---

        if 'excel_data' in st.session_state and not st.session_state.excel_data.empty:
            try:
                # 불일치 데이터 생성
                new_mismatch_result = data_analyzer.find_mismatches(st.session_state.excel_data)
                if new_mismatch_result["status"] == "success":
                    new_mismatch_data = new_mismatch_result["data"]

                    # 제외할 물품코드 제거 (하드코딩)
                    excluded_item_codes = [
                        'L505001', 'L505002', 'L505003', 'L505004', 'L505005', 'L505006', 'L505007', 
                        'L505008', 'L505009', 'L505010', 'L505011', 'L505012', 'L505013', 'L505014',
                        'L605001', 'L605002', 'L605003', 'L605004', 'L605005', 'L605006'
                    ]
                    if not new_mismatch_data.empty and '물품코드' in new_mismatch_data.columns:
                        new_mismatch_data = new_mismatch_data[
                            ~new_mismatch_data['물품코드'].astype(str).isin(excluded_item_codes)
                        ]

                    # 완료 처리 로그 필터링 (세션 상태 사용)
                    completion_logs = st.session_state.get('completion_logs', [])
                    if not new_mismatch_data.empty and completion_logs:
                        new_mismatch_data = filter_completed_items(new_mismatch_data, completion_logs)

                    st.session_state.mismatch_data = new_mismatch_data.reset_index(drop=True)
                    
                    # 통합 파일 업데이트 제거 - 사용자가 부서별 통계 탭에서 직접 병합 버튼을 눌러야 함
                    # 날짜별 S3 저장은 이미 위에서 완료됨
                    logger.info("날짜별 S3 저장 완료. 통합 파일 업데이트는 부서별 통계 탭에서 수동으로 수행하세요.")

                else:
                    st.session_state.mismatch_data = pd.DataFrame()
                    st.warning(f"새 불일치 데이터 계산 실패: {new_mismatch_result['message']}")
            except Exception as e:
                st.session_state.mismatch_data = pd.DataFrame()
                logger.error(f"불일치 데이터 계산/저장 중 오류: {e}", exc_info=True)
                st.error("불일치 데이터를 계산하거나 저장하는 중 오류가 발생했습니다.")
        else:
            st.session_state.mismatch_data = pd.DataFrame()
            logger.info("엑셀 데이터가 없어 불일치 데이터를 계산하지 않습니다.")
        # -------------------------------------
        
        # --- 7. 메타데이터 업데이트 ---
        final_processed_dates = sorted(list(processed_dates))
        for date_str in final_processed_dates:
            try:
                metadata_result = s3_handler.load_metadata(date_str)
                if metadata_result["status"] == "success":
                    metadata = metadata_result["data"]
                else:
                    metadata = {} # 기존 메타데이터 없음
                
                # 엑셀 정보 업데이트 (누적 파일 기준)
                # 엑셀 정보 업데이트 (누적 파일 기준)
                metadata["excel_key"] = cumulative_excel_key
                metadata["excel_hash"] = cumulative_excel_hash # 위에서 계산한 누적 해시
                metadata["excel_processed_files"] = newly_processed_excel_files # 이번 실행에서 처리한 파일 목록tr]
                if date_str in st.session_state.ocr_results_by_date:
                    ocr_data = st.session_state.ocr_results_by_date[date_str]
                    metadata["pdf_filename"] = metadata.get("pdf_filename", "N/A") # 이전 값 유지 시도
                    metadata["ocr_pages"] = len(ocr_data.get("ocr_text", []))
                    metadata["departments_with_pages"] = ocr_data.get("departments_with_pages", [])
                
                metadata["processed_date"] = datetime.now().isoformat()
                s3_handler.save_metadata(date_str, metadata)
                logger.debug(f"메타데이터 업데이트 완료: {date_str}")
            except Exception as e:
                logger.error(f"메타데이터 업데이트 실패 ({date_str}): {e}", exc_info=True)
                st.warning(f"{date_str} 날짜의 메타데이터 업데이트 중 오류 발생")
        # -------------------------------------
        
        # --- 8. 사용 가능한 날짜 목록 업데이트 및 마무리 --- 
        st.session_state.available_dates = final_processed_dates
        # logger.info(f"최종 처리된 날짜 목록 업데이트: {len(st.session_state.available_dates)}개") # 중복 로그 제거

        # PDF 처리 결과 디버깅 로그 추가
        for date in final_processed_dates:
            if date in st.session_state.pdf_paths_by_date:
                logger.info(f"날짜 {date}의 PDF 경로: {st.session_state.pdf_paths_by_date[date]}")
            if date in st.session_state.departments_with_pages_by_date:
                dept_pages = st.session_state.departments_with_pages_by_date[date]
                # logger.info(f"날짜 {date}의 부서-페이지 정보: {len(dept_pages)}개 항목") # 중복 로그 제거
            else:
                logger.warning(f"날짜 {date}의 부서-페이지 정보가 없습니다.")

        if final_processed_dates:
            # 가장 최근 날짜로 선택 업데이트 (선택적)
            new_date = max(final_processed_dates)
            st.session_state.selected_date = new_date
            # logger.info(f"가장 최근 처리 날짜로 선택 업데이트: {new_date}") # 중복 로그 제거

        st.success("모든 파일 처리가 완료되었습니다!")
        st.session_state.force_reload_mismatch = True
    except Exception as e:
        st.error(f"파일 처리 중 예기치 않은 오류가 발생했습니다: {str(e)}")
        logger.exception("파일 처리 중 최상위 오류 발생")
        # ---------------------------------------------

    except Exception as e:
        st.error(f"파일 처리 중 예기치 않은 오류가 발생했습니다: {str(e)}")
        logger.exception("파일 처리 중 최상위 오류 발생")

@st.cache_data(ttl=3600)
def get_pdf_preview_image_from_s3(file_key):
    s3_handler = S3Handler()
    result = s3_handler.download_file(file_key)
    if result["status"] == "success":
        return result["data"]
    return None

# 불일치 리스트 탭 표시 함수
def display_mismatch_tab(): # selected_date 인자 제거
    """날짜별 작업 탭을 표시합니다."""
    try:
        # --- 날짜 선택 UI (탭 내부) ---
        st.header("날짜별 작업 상세 조회")
        if not st.session_state.get("available_dates"):
            st.warning("사이드바에서 작업 기간을 먼저 설정하고 파일을 처리해주세요.")
            return

        selected_date_in_tab = st.selectbox(
            "작업할 날짜를 선택하세요:",
            st.session_state.available_dates,
            key="selected_date_in_mismatch_tab"
        )

        if not selected_date_in_tab:
            st.info("날짜를 선택해주세요.")
            return
        
        # 선택된 날짜에 대한 데이터 로드 (main에서 이동)
        if selected_date_in_tab not in st.session_state.get('loaded_dates', set()):
            with st.spinner(f"{selected_date_in_tab} 날짜의 데이터를 불러오는 중..."):
                result = load_data_for_date(selected_date_in_tab)
                if result.get("status") == "success":
                    if 'loaded_dates' not in st.session_state:
                        st.session_state.loaded_dates = set()
                    st.session_state.loaded_dates.add(selected_date_in_tab)
                    # 성공 메시지 제거하여 중복 새로고침 방지
                else:
                    st.warning(f"{selected_date_in_tab} 날짜 데이터 로드 실패: {result.get('message')}")
        
        # PDF 존재 여부 확인 (S3에서 직접 확인)
        s3_handler = S3Handler()
        
        # 1. 세션 상태에서 먼저 확인
        pdf_exists_in_session = selected_date_in_tab in st.session_state.get('pdf_paths_by_date', {})
        
        # 2. 세션에 없으면 S3 메타데이터에서 확인
        pdf_exists_in_s3 = False
        if not pdf_exists_in_session:
            metadata_result = s3_handler.load_metadata(selected_date_in_tab)
            if metadata_result["status"] == "success":
                metadata = metadata_result["data"]
                if "pdf_key" in metadata:
                    # 세션에 PDF 경로 저장
                    st.session_state.pdf_paths_by_date[selected_date_in_tab] = metadata["pdf_key"]
                    pdf_exists_in_s3 = True
                    
                    # 부서-페이지 정보도 세션에 저장
                    if "departments_with_pages" in metadata:
                        st.session_state.departments_with_pages_by_date[selected_date_in_tab] = metadata["departments_with_pages"]
        
        # PDF 존재 여부 표시
        if pdf_exists_in_session or pdf_exists_in_s3:
            st.success(f"선택된 날짜({selected_date_in_tab})에 PDF 파일이 있습니다.")
        else:
            st.warning(f"선택된 날짜({selected_date_in_tab})에 PDF 파일이 없습니다.")
        # --- 날짜 선택 UI 끝 ---
        
        if 'mismatch_data' not in st.session_state or st.session_state.mismatch_data.empty:
            st.info("처리된 불일치 데이터가 없습니다.")
            return
            
        # 2) 완료 처리된 항목 필터링 (세션 상태 사용)
        completion_logs = st.session_state.get('completion_logs', [])
        if completion_logs:
            filtered_mismatch_data = filter_completed_items(st.session_state.mismatch_data, completion_logs)
        else:
            filtered_mismatch_data = st.session_state.mismatch_data
            
        # 3) 날짜별 필터링
        # 날짜 컬럼이 문자열인 경우 datetime으로 변환
        if not pd.api.types.is_datetime64_any_dtype(filtered_mismatch_data['날짜']):
            filtered_mismatch_data['날짜'] = pd.to_datetime(filtered_mismatch_data['날짜'], format='%Y-%m-%d', errors='coerce')
            
        df_date = filtered_mismatch_data[
            filtered_mismatch_data['날짜'].dt.strftime('%Y-%m-%d') == selected_date_in_tab # selected_date_in_tab 사용
        ].copy()
        
        if df_date.empty:
            st.info(f"선택된 날짜({selected_date_in_tab})에 해당하는 불일치 데이터가 없습니다.") # selected_date_in_tab 사용
            # 이 경우에도 특정 부서 탭으로 바로 넘어갈 수 있으므로, 전체 탭에 대한 처리는 계속 진행
            # return # 여기서 리턴하면 전체 탭 및 부서별 탭이 아예 안 나옴
            
        dept_result = get_unique_departments(df_date)
        if dept_result["status"] == "error":
            st.error(dept_result["message"])
            return
            
        dept_options = dept_result["data"]
        
        # PDF의 부서 목록도 가져오기
        pdf_depts = set()
        if selected_date_in_tab in st.session_state.get('departments_with_pages_by_date', {}):
            dept_page_tuples = st.session_state.departments_with_pages_by_date[selected_date_in_tab]
            pdf_depts = {dept for dept, page in dept_page_tuples}
        
        # 불일치 데이터가 있는 부서와 PDF에만 있는 부서를 모두 포함
        all_dept_options = sorted(list(set(dept_options) | pdf_depts))
        
        # 5) 부서별 서브탭 생성 (모든 부서 포함)
        dept_tabs = st.tabs(["전체"] + all_dept_options)
        
        # 전체 탭 (일괄 처리 + 부서 비교 전용)
        with dept_tabs[0]:
            st.subheader("📋 선택 항목 관리")
            
            # 선택 상태 요약 표시 (자동 갱신)
            selected_count_by_dept = {}
            total_selected = 0
            
            # 각 부서별로 선택된 항목 수 계산 (통합된 키 사용)
            for dept in dept_options:
                count = 0
                dept_data = df_date[df_date['부서명'] == dept]
                for idx, row in dept_data.iterrows():
                    try:
                        date_val = pd.to_datetime(row.get('날짜', 'N/A')).strftime('%Y-%m-%d')
                    except:
                        date_val = str(row.get('날짜', 'N/A'))
                    dept_key_val = str(row.get('부서명', 'N/A'))
                    code_key_val = str(row.get('물품코드', 'N/A'))
                    # 부서별 탭과 동일한 키 형식 사용 (부서 접미사 제거)
                    state_key = f"sel_{date_val}_{dept_key_val}_{code_key_val}"
                    
                    if st.session_state.get(state_key, False):
                        count += 1
                        total_selected += 1
                        
                selected_count_by_dept[dept] = count
            
            # 선택 저장 상태 확인
            saved_selections = st.session_state.get('saved_selections', {})
            saved_count = sum(saved_selections.values())
            
            col1, col2, col3 = st.columns([1, 1, 1])
            with col1:
                st.metric("총 선택 항목", f"{total_selected}개")
            with col2:
                st.metric("저장된 선택", f"{saved_count}개")
            with col3:
                if total_selected > 0:
                    if saved_count == 0:
                        st.warning("⚠️ 선택 저장 필요")
                    elif saved_count < total_selected:
                        st.info("💡 일부 저장됨")
                    else:
                        st.success("✅ 모두 저장됨")
            
            # 부서별 선택 상태 표시
            if total_selected > 0 or saved_count > 0:
                dept_summary = []
                for dept, count in selected_count_by_dept.items():
                    if count > 0:
                        # 해당 부서의 저장 상태 확인
                        dept_key = f"{selected_date_in_tab}_{dept}"
                        saved_for_dept = saved_selections.get(dept_key, 0)
                        if saved_for_dept > 0:
                            dept_summary.append(f"{dept}: {count}개 (저장됨: {saved_for_dept}개)")
                        else:
                            dept_summary.append(f"{dept}: {count}개 (미저장)")
                
                if dept_summary:
                    st.info("부서별 선택: " + ", ".join(dept_summary))
                
                # 일괄 완료 처리 버튼 (저장된 선택이 있을 때만 활성화)
                if st.button("🚀 모든 부서 선택 항목 일괄 완료 처리", 
                           type="primary", 
                           key="batch_complete_all",
                           disabled=(saved_count == 0),
                           help="각 부서에서 '선택 저장'을 먼저 눌러주세요" if saved_count == 0 else "저장된 선택 항목을 일괄 완료 처리합니다"):
                    with st.spinner("일괄 완료 처리 중... (S3 저장 및 통합 작업 수행)"):
                        all_completed_items = []
                        all_indices_to_remove = []
                        
                        # 모든 부서의 선택된 항목 수집 (통합된 키 사용)
                    for dept in dept_options:
                        dept_data = df_date[df_date['부서명'] == dept]
                        for idx, row in dept_data.iterrows():
                            try:
                                date_val = pd.to_datetime(row.get('날짜', 'N/A')).strftime('%Y-%m-%d')
                            except:
                                date_val = str(row.get('날짜', 'N/A'))
                            dept_key_val = str(row.get('부서명', 'N/A'))
                            code_key_val = str(row.get('물품코드', 'N/A'))
                            # 부서별 탭과 동일한 키 형식 사용 (부서 접미사 제거)
                            state_key = f"sel_{date_val}_{dept_key_val}_{code_key_val}"
                            
                            if st.session_state.get(state_key, False):
                                original_idx = row.get('original_index', idx)
                                all_indices_to_remove.append(original_idx)
                                all_completed_items.append({
                                    '날짜': date_val,
                                    '부서명': dept_key_val,
                                    '물품코드': code_key_val,
                                    '물품명': row.get('물품명', 'N/A'),
                                    '청구량': row.get('청구량', 0),
                                    '수령량': row.get('수령량', 0),
                                    '차이': row.get('차이', 0),
                                    '누락': row.get('누락', ''),
                                    '처리시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    'original_index': original_idx
                                })
                                # 선택 상태 초기화
                                if state_key in st.session_state:
                                    del st.session_state[state_key]
                    
                    # 일괄 처리 실행
                    if all_indices_to_remove:
                        # mismatch_data에서 제거
                        st.session_state.mismatch_data = st.session_state.mismatch_data.drop(
                            all_indices_to_remove
                        ).reset_index(drop=True)
                        
                        # 전산누락 저장 시에만 필요한 자동 통합 작업 제거
                        # 사용자가 명시적으로 부서별 통계 탭에서 병합 버튼을 누르도록 유도
                        
                        # 완료 처리 로그 저장
                        if all_completed_items:
                            log_result = s3_handler.save_completion_log(all_completed_items)
                            if log_result["status"] != "success":
                                st.warning("완료 처리 로그 저장에 실패했습니다.")
                            
                            # 세션 상태에도 완료 처리 로그 추가
                            if 'completion_logs' not in st.session_state:
                                st.session_state.completion_logs = []
                            st.session_state.completion_logs.extend(all_completed_items)
                            
                            # 중복 제거
                            if st.session_state.completion_logs:
                                temp_df = pd.DataFrame(st.session_state.completion_logs)
                                if '처리시간' in temp_df.columns:
                                    temp_df['처리시간'] = pd.to_datetime(temp_df['처리시간'])
                                    temp_df = temp_df.sort_values('처리시간', ascending=False)
                                    temp_df = temp_df.drop_duplicates(subset=['날짜', '부서명', '물품코드'], keep='first')
                                    st.session_state.completion_logs = temp_df.to_dict('records')
                        
                        # 선택 저장 플래그 모두 정리
                        if 'saved_selections' in st.session_state:
                            st.session_state.saved_selections.clear()
                    
                    if all_indices_to_remove:
                        st.success(f"✅ 총 {len(all_indices_to_remove)}개 항목이 일괄 완료 처리되었습니다! (날짜별 저장 완료)")
                        st.info("💡 부서별 통계를 보려면 '날짜별 작업 내용 병합' 버튼을 눌러주세요.")
                        st.balloons()
                    else:
                        st.warning("선택된 항목이 없습니다.")
            else:
                st.info("💡 각 부서 탭에서 완료 처리할 항목을 선택해주세요.")
            
            st.markdown("---")

            
            # 엑셀과 PDF의 부서 비교
            st.subheader("📋 PDF & 엑셀 부서 비교")
            try:
                # 엑셀의 부서 목록
                excel_depts = set()
                if 'excel_data' in st.session_state and not st.session_state.excel_data.empty:
                    # 날짜 형식 변환
                    if not pd.api.types.is_datetime64_any_dtype(st.session_state.excel_data['날짜']):
                        st.session_state.excel_data['날짜'] = pd.to_datetime(st.session_state.excel_data['날짜'], format='%Y-%m-%d', errors='coerce')
                    
                    excel_date_data = st.session_state.excel_data[
                        st.session_state.excel_data['날짜'].dt.strftime('%Y-%m-%d') == selected_date_in_tab
                    ]
                    excel_depts = set(excel_date_data['부서명'].unique())
                
                # PDF의 부서 목록
                pdf_depts = set()
                if selected_date_in_tab in st.session_state.get('departments_with_pages_by_date', {}):
                    dept_page_tuples = st.session_state.departments_with_pages_by_date[selected_date_in_tab]
                    pdf_depts = {dept for dept, page in dept_page_tuples}
                
                # 부서 비교 결과 표시
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("엑셀 부서 수", len(excel_depts))
                with col2:
                    st.metric("PDF 부서 수", len(pdf_depts))
                with col3:
                    common_depts = excel_depts & pdf_depts
                    st.metric("공통 부서 수", len(common_depts))
                
                # PDF에만 있는 부서 (누락된 부서)
                pdf_only_depts = pdf_depts - excel_depts
                if pdf_only_depts:
                    st.warning(f"⚠️ PDF에만 있는 부서 ({len(pdf_only_depts)}개)")
                    st.write("**누락된 부서 목록:**", ", ".join(sorted(pdf_only_depts)))
                    
                    # 누락된 부서의 PDF 미리보기 표시
                    st.subheader("📄 누락된 부서 PDF 미리보기")
                    for dept in sorted(pdf_only_depts):
                        with st.expander(f"📁 {dept} 부서 PDF 미리보기"):
                            dept_pages = get_department_pages(selected_date_in_tab, dept)
                            if dept_pages:
                                # PDF 원본 다운로드
                                pdf_key = st.session_state.pdf_paths_by_date.get(selected_date_in_tab)
                                if pdf_key:
                                    pdf_result = s3_handler.download_file(pdf_key)
                                    if pdf_result["status"] == "success":
                                        pdf_bytes = pdf_result["data"]
                                        
                                        # 부서의 각 페이지 미리보기 표시
                                        cols = st.columns(min(2, len(dept_pages)))
                                        for i, page_num in enumerate(dept_pages[:2]):  # 최대 2개 페이지만 표시
                                            with cols[i % 2]:
                                                img = extract_pdf_preview(
                                                    io.BytesIO(pdf_bytes), 
                                                    page_num-1, 
                                                    dpi=120, 
                                                    thumbnail_size=(700, 1000)
                                                )
                                                if img:
                                                    st.image(img, caption=f"페이지 {page_num}")
                                        
                                        if len(dept_pages) > 2:
                                            st.info(f"총 {len(dept_pages)}개 페이지 중 2개만 표시됨")
                            else:
                                st.info("해당 부서의 페이지 정보를 찾을 수 없습니다.")
                
                # 엑셀에만 있는 부서
                excel_only_depts = excel_depts - pdf_depts
                if excel_only_depts:
                    st.info(f"ℹ️ 엑셀에만 있는 부서 ({len(excel_only_depts)}개): {', '.join(sorted(excel_only_depts))}")
                
                if not pdf_only_depts and not excel_only_depts:
                    st.success("✅ 모든 부서가 엑셀과 PDF에 일치합니다!")
                    
            except Exception as e:
                logger.error(f"부서 비교 중 오류 발생: {e}", exc_info=True)
                st.error("부서 비교 중 오류가 발생했습니다.")
        
        # 각 부서별 탭
        for i, dept in enumerate(all_dept_options, 1):
            with dept_tabs[i]:
                # 불일치 데이터가 있는 부서인지 확인
                if dept in dept_options:
                    df_filtered_dept = df_date[df_date['부서명'] == dept].copy()
                else:
                    # PDF에만 있는 부서 (불일치 데이터 없음)
                    df_filtered_dept = pd.DataFrame()
                    st.info(f"ℹ️ '{dept}' 부서의 전산 누락 품목을 확인할 수 있습니다.")
                    st.warning("💡 아래에서 PDF 품목을 확인하고 필요시 전산누락으로 저장할 수 있습니다.")
                    st.caption("완료 처리로 인해 불일치가 모두 해결된 경우에도 이 메시지가 나타날 수 있습니다.")
                
                st.subheader("PDF & 엑셀 품목 비교")
                try:
                    excel_items_result = get_excel_items(selected_date_in_tab, dept) # selected_date_in_tab 사용
                    if excel_items_result["status"] == "success":
                        excel_data_from_func = excel_items_result["data"]
                        
                        if isinstance(excel_data_from_func, pd.DataFrame):
                            excel_df = excel_data_from_func
                        elif isinstance(excel_data_from_func, list):
                            logger.warning(f"get_excel_items가 list를 반환 (부서: {dept}). DataFrame 변환 시도.")
                            try:
                                if excel_data_from_func and isinstance(excel_data_from_func[0], str):
                                    excel_df = pd.DataFrame(excel_data_from_func, columns=['물품코드'])
                                    logger.info(f"단순 list를 '물품코드' 컬럼 DataFrame으로 변환 (부서: {dept})")
                                else:
                                    excel_df = pd.DataFrame(excel_data_from_func)
                                
                                if excel_df.empty and excel_data_from_func:
                                    logger.warning(f"리스트로부터 빈 DataFrame 생성 (부서: {dept}). 예상 컬럼으로 재생성.")
                                    excel_df = pd.DataFrame(columns=['물품코드', '물품명', '청구량'])
                                elif not excel_df.empty and '물품코드' not in excel_df.columns:
                                    logger.warning(f"생성된 DataFrame에 '물품코드' 컬럼 없음 (부서: {dept}). 예상 컬럼으로 재생성.")
                                    excel_df = pd.DataFrame(columns=['물품코드', '물품명', '청구량'])

                            except Exception as e:
                                logger.error(f"list를 DataFrame으로 변환 중 오류 (부서: {dept}): {e}")
                                excel_df = pd.DataFrame(columns=['물품코드', '물품명', '청구량'])
                        else:
                            logger.warning(f"get_excel_items가 예상치 않은 타입({type(excel_data_from_func)})을 반환 (부서: {dept}). 빈 DataFrame 사용.")
                            excel_df = pd.DataFrame(columns=['물품코드', '물품명', '청구량'])
                        
                        pdf_item_set = get_department_items(selected_date_in_tab, dept) # selected_date_in_tab 사용
                        
                        if pdf_item_set: 
                            pdf_df = pd.DataFrame({"물품코드": list(pdf_item_set)})
                        else: 
                            pdf_df = pd.DataFrame(columns=["물품코드"])

                        if not pdf_df.empty:
                            if not excel_df.empty and '물품코드' in excel_df.columns: 
                                pdf_only_codes = pdf_df[~pdf_df['물품코드'].isin(excel_df['물품코드'])]['물품코드'].tolist()
                            else: 
                                pdf_only_codes = pdf_df['물품코드'].tolist()
                            
                            if pdf_only_codes:
                                missing_from_excel_items = []
                                item_db = st.session_state.get("item_db", {}) 
                                for code in pdf_only_codes:
                                    item_name = item_db.get(code, "알 수 없는 물품") 
                                    pdf_quantity = 1 
                                    missing_item_details = {
                                        '날짜': selected_date_in_tab, '부서명': dept, '물품코드': code, # selected_date_in_tab
                                        '물품명': item_name, '청구량': 0, '수령량': pdf_quantity, 
                                        '차이': pdf_quantity, '누락': '전산누락' 
                                    }
                                    missing_from_excel_items.append(missing_item_details)
                                if missing_from_excel_items:
                                    missing_df = pd.DataFrame(missing_from_excel_items)
                                    # df_filtered_dept = pd.concat([df_filtered_dept, missing_df], ignore_index=True).drop_duplicates(
                                    #     subset=['날짜', '부서명', '물품코드'], keep='last'
                                    # ) # 바로 합치지 않음

                                    # 사용자에게 감지된 누락 항목 보여주기
                                    with st.expander(f"자동 감지된 전산누락 후보 ({len(missing_df)}개) - 검토 후 저장하세요", expanded=True):
                                        st.info("아래 목록은 PDF 인수증에는 있지만 엑셀 데이터에는 없는 품목들입니다. 검토 후 '전산누락 저장' 버튼을 눌러야 통계에 반영됩니다.")
                                        st.dataframe(missing_df[['날짜', '부서명', '물품코드', '물품명', '수령량']].rename(columns={'수령량':'PDF수량'}))

                                        form_key_missing_save = f"form_missing_save_{selected_date_in_tab}_{dept}"
                                        with st.form(key=form_key_missing_save):
                                            save_detected_missing_button = st.form_submit_button("✅ 위 전산누락 후보 저장")

                                            if save_detected_missing_button:
                                                try:
                                                    # 캐시 클리어 추가
                                                    st.cache_data.clear()
                                                    
                                                    # 전산누락 저장 전 디버깅 정보
                                                    logger.info(f"전산누락 저장 시작 - 날짜: {selected_date_in_tab}, 부서: {dept}, 항목 수: {len(missing_df)}")
                                                    logger.info(f"전산누락 데이터 샘플: {missing_df[['날짜', '부서명', '물품코드', '누락']].head().to_dict('records')}")
                                                    
                                                    s3_handler = S3Handler()
                                                    result = s3_handler.save_missing_items_by_date(missing_df, date_str=selected_date_in_tab)
                                                    
                                                    logger.info(f"전산누락 S3 저장 결과: {result['status']} - {result.get('message', '')}")
                                                    
                                                    if result["status"] == "success":
                                                        # 세션 상태의 mismatch_data도 업데이트
                                                        
                                                        if 'mismatch_data' not in st.session_state:
                                                            st.session_state.mismatch_data = pd.DataFrame()
                                                        
                                                        # 기존 데이터와 새 전산누락 데이터 병합
                                                        combined_session = pd.concat([st.session_state.mismatch_data, missing_df], ignore_index=True)
                                                        # 중복 제거
                                                        combined_session = combined_session.drop_duplicates(subset=['날짜', '부서명', '물품코드'], keep='last')
                                                        st.session_state.mismatch_data = combined_session
                                                        
                                                        # 강제 새로고침 플래그 설정 (부서별 통계 탭 자동 업데이트)
                                                        st.session_state.force_refresh = True
                                                        
                                                        st.success(f"{len(missing_df)}개 전산누락 항목이 저장되었습니다!")
                                                        st.info("💡 부서별 통계 탭에서 '날짜별 작업 내용 병합' 버튼을 눌러 확인하세요.")
                                                        
                                                        # 페이지 새로고침으로 즉시 반영
                                                        # time.sleep(1)  # 잠시 대기 후 새로고침
                                                        # st.rerun()
                                                except Exception as e:
                                                    st.error(f"전산누락 저장 중 오류: {e}")
                                                    logger.error(f"전산누락 저장 중 오류: {e}", exc_info=True)
                           
                    else: 
                        st.error(f"'{dept}' 부서의 엑셀 데이터를 불러오는데 실패했습니다: {excel_items_result.get('message', '알 수 없는 오류')}")
                        excel_df = pd.DataFrame(columns=['물품코드', '물품명', '청구량']) 
                
                except Exception as e:
                    logger.error(f"PDF & 엑셀 품목 비교 중 오류 발생 ({dept}): {e}", exc_info=True)
                    st.error("PDF & 엑셀 품목 비교 중 오류가 발생했습니다.")
                    # 이 경우에도 excel_df가 정의되지 않았을 수 있으므로, 또는 try 블록 시작 전에 초기화 필요
                    # df_filtered는 이미 이 try블록 외부에서 해당 dept로 필터링된 데이터로 존재함

                # 최종적으로 df_filtered_dept를 사용해 display_mismatch_content 호출
                # 불일치 데이터가 없어도 전산누락 확인을 위해 항상 호출
                display_mismatch_content(df_filtered_dept, selected_date_in_tab, dept, s3_handler)
                # display_pdf_section 중복 호출 제거 - display_mismatch_content 내부에서 이미 호출됨
    except Exception as e:
        logger.error(f"display_mismatch_tab 오류: {e}", exc_info=True)
        st.error(f"데이터 표시 중 오류가 발생했습니다: {e}")

def get_excel_items(date_str, dept_name):
    """
    특정 날짜와 부서의 엑셀 품목 정보(물품코드, 물품명, 청구량)를 DataFrame으로 반환합니다.
    (필요 컬럼 없을 때도 에러 안나고, 항상 컬럼명 유지)
    """
    try:
        if 'excel_data' in st.session_state and not st.session_state.excel_data.empty:
            dept_excel_data = st.session_state.excel_data[
                (st.session_state.excel_data['날짜'] == date_str) &
                (st.session_state.excel_data['부서명'] == dept_name)
            ].copy()
            # 기본 반환 컬럼
            required_cols = ['물품코드', '물품명', '청구량']
            # 실제 있는 컬럼만 추출, 없으면 빈 DF
            if not dept_excel_data.empty:
                available_cols = [col for col in required_cols if col in dept_excel_data.columns]
                if '물품코드' not in available_cols:
                    return {"status": "error", "message": "필수 컬럼 '물품코드'가 없습니다."}
                return {"status": "success", "data": dept_excel_data[available_cols]}
            else:
                return {"status": "success", "data": pd.DataFrame(columns=required_cols)}
        return {"status": "error", "message": "세션에 엑셀 데이터가 없습니다."}
    except Exception as e:
        logger.error(f"get_excel_items 오류: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


def display_mismatch_content(df_filtered, selected_date, sel_dept, s3_handler):
    """불일치 데이터 표시 내용을 처리하는 함수"""
    try:
        # original_index 컬럼 추가
        if 'original_index' not in df_filtered.columns:
            df_filtered['original_index'] = df_filtered.index
            
        # 불일치 데이터가 없어도 계속 진행 (전산누락 확인을 위해)
        if not df_filtered.empty:
            st.markdown("**완료 처리할 항목을 선택하세요.**")
            
            # 선택 저장 form (체크박스 포함)
            form_key_selection = f"selection_form_{selected_date}_{sel_dept}"
            with st.form(key=form_key_selection):
                # '누락' 열을 포함하여 widths 리스트 수정 (총 9개 열)
                widths = [0.5, 1.2, 0.8, 0.8, 2.5, 0.7, 0.7, 0.7, 1] 
                
                # 헤더 표시
                header_cols = st.columns(widths)
                column_names = ["선택", "날짜", "부서명", "물품코드", "물품명", "청구량", "수령량", "차이", "누락"]
                for i, name in enumerate(column_names):
                    if i < len(header_cols):
                        header_cols[i].markdown(f"**{name}**")
                
                # 체크박스와 데이터 표시 (form 안에서)
                selected_items = []
                for idx, row in df_filtered.iterrows():
                    try:
                        date_val = pd.to_datetime(row.get('날짜', 'N/A')).strftime('%Y-%m-%d')
                    except:
                        date_val = str(row.get('날짜', 'N/A'))
                        
                    dept_key_val = str(row.get('부서명', 'N/A'))
                    code_key_val = str(row.get('물품코드', 'N/A'))
                    # 전체 탭과 동일한 키 형식 사용 (부서 접미사 제거)
                    state_key = f"sel_{date_val}_{dept_key_val}_{code_key_val}"
                    
                    cols = st.columns(widths)
                    # label을 고유하게 만들고 숨김 처리
                    checkbox_label = f"select_{state_key}"
                    is_selected = cols[0].checkbox(
                        label=checkbox_label, 
                        key=f"{form_key_selection}_{state_key}",  # form 내부 고유 키 사용
                        value=st.session_state.get(state_key, False),
                        label_visibility="collapsed"
                    )
                    
                    if is_selected:
                        selected_items.append((state_key, row))
                    
                    try:
                        # 각 컬럼에 해당하는 값을 안전하게 가져와서 표시
                        col_values = [
                            date_val,
                            dept_key_val,
                            code_key_val,
                            str(row.get('물품명', row.get('품목', 'N/A'))),
                            str(row.get('청구량', 'N/A')),
                            str(row.get('수령량', 'N/A')),
                            str(row.get('차이', 'N/A')),
                            str(row.get('누락', ''))
                        ]
                        for i, value in enumerate(col_values):
                            if (i + 1) < len(cols):
                                cols[i+1].write(value)
                    except Exception as row_err:
                        logger.error(f"불일치 리스트 행 값 표시 오류 (인덱스: {idx}, 데이터: {row.to_dict()}): {row_err}")
                        # 오류 발생 시 대체 텍스트 표시 (선택 열 제외)
                        for i in range(1, len(cols)):
                            cols[i].write("-")
                
                st.markdown("---")
                col1, col2 = st.columns([1, 1])
                with col1:
                    save_selection_button = st.form_submit_button("💾 선택 저장", type="secondary", 
                                                                help="체크박스 선택을 세션에 저장합니다 (UI 새로고침 없음)")
                with col2:
                    immediate_complete_button = st.form_submit_button("✅ 즉시 완료 처리", type="primary",
                                                                    help="선택한 항목을 바로 완료 처리합니다 (UI 새로고침 발생)")

                # 선택 저장 처리 (UI 새로고침 없음, S3 작업 없음) - 최적화됨
                if save_selection_button:
                    # 1. 선택된 항목들의 키 집합 생성 (빠른 검색용)
                    selected_keys = {state_key for state_key, row in selected_items}
                    
                    # 2. 선택된 항목들을 True로 설정
                    for state_key in selected_keys:
                        st.session_state[state_key] = True
                    
                    # 3. 선택되지 않은 항목들을 False로 설정 (최적화)
                    # 날짜 변환을 한 번만 수행
                    try:
                        date_val = pd.to_datetime(df_filtered['날짜'].iloc[0]).strftime('%Y-%m-%d')
                    except:
                        date_val = str(df_filtered['날짜'].iloc[0])
                    
                    dept_key_val = str(df_filtered['부서명'].iloc[0])  # 같은 부서이므로 첫 번째 값 사용
                    
                    # 벡터화된 키 생성
                    code_values = df_filtered['물품코드'].astype(str)
                    all_keys = {f"sel_{date_val}_{dept_key_val}_{code}" for code in code_values}
                    
                    # 선택되지 않은 키들만 False로 설정
                    unselected_keys = all_keys - selected_keys
                    for key in unselected_keys:
                        st.session_state[key] = False
                    
                    # 4. 선택 저장 완료 플래그 설정 (전체 탭에서 확인용)
                    if 'saved_selections' not in st.session_state:
                        st.session_state.saved_selections = {}
                    st.session_state.saved_selections[f"{selected_date}_{sel_dept}"] = len(selected_items)
                    
                    st.success(f"✅ {len(selected_items)}개 항목 선택이 저장되었습니다. 전체 탭에서 일괄 처리하세요.")
                    st.info("💡 이 작업은 세션에만 저장되며 S3 작업이 없어 빠릅니다.")

                # 즉시 완료 처리 (S3 작업 포함, 시간 소요)
                if immediate_complete_button:
                    if selected_items:
                        with st.spinner("완료 처리 중... (S3 저장 및 통합 작업 수행)"):
                            items_to_remove_keys = []
                            items_to_remove_indices = []
                            completed_items = []

                        for state_key, row in selected_items:
                            try:
                                date_k = pd.to_datetime(row.get('날짜', 'N/A')).strftime('%Y-%m-%d')
                            except:
                                date_k = str(row.get('날짜', 'N/A'))
                                
                            dept_k = str(row.get('부서명', 'N/A'))
                            code_k = str(row.get('물품코드', 'N/A'))
                            original_idx = row['original_index']
                            
                            items_to_remove_keys.append(state_key)
                            items_to_remove_indices.append(original_idx)
                            completed_items.append({
                                '날짜': date_k,
                                '부서명': dept_k,
                                '물품코드': code_k,
                                '물품명': row.get('물품명', 'N/A'),
                                '청구량': row.get('청구량', 0),
                                '수령량': row.get('수령량', 0),
                                '차이': row.get('차이', 0),
                                '누락': row.get('누락', ''),
                                '처리시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'original_index': original_idx
                            })

                        if items_to_remove_indices:
                            st.session_state.mismatch_data = st.session_state.mismatch_data.drop(items_to_remove_indices).reset_index(drop=True)
                                                   
                                   
                            
                            if completed_items:
                                # S3에 저장
                                log_result = s3_handler.save_completion_log(completed_items)
                                if log_result["status"] != "success":
                                    st.warning("완료 처리 로그 저장에 실패했습니다.")
                                
                                # 세션 상태에도 완료 처리 로그 추가
                                if 'completion_logs' not in st.session_state:
                                    st.session_state.completion_logs = []
                                st.session_state.completion_logs.extend(completed_items)
                                
                                # 중복 제거
                                if st.session_state.completion_logs:
                                    temp_df = pd.DataFrame(st.session_state.completion_logs)
                                    if '처리시간' in temp_df.columns:
                                        temp_df['처리시간'] = pd.to_datetime(temp_df['처리시간'])
                                        temp_df = temp_df.sort_values('처리시간', ascending=False)
                                        temp_df = temp_df.drop_duplicates(subset=['날짜', '부서명', '물품코드'], keep='first')
                                        st.session_state.completion_logs = temp_df.to_dict('records')
                            
                        # 세션 정리 (완료 처리된 항목들)
                        for key in items_to_remove_keys:
                            if key in st.session_state:
                                del st.session_state[key]
                        
                        # 선택 저장 플래그도 정리
                        if 'saved_selections' in st.session_state:
                            key_to_remove = f"{selected_date}_{sel_dept}"
                            if key_to_remove in st.session_state.saved_selections:
                                del st.session_state.saved_selections[key_to_remove]
                        
                    st.success(f"✅ {len(items_to_remove_indices)}개 항목이 완료 처리되었습니다. (날짜별 저장 완료)")
                    st.info("💡 부서별 통계를 보려면 '날짜별 작업 내용 병합' 버튼을 눌러주세요.")
                else:
                    st.warning("완료 처리할 항목을 선택하세요.")

            # PDF 섹션 표시 - 불일치 데이터 유무와 관계없이 항상 표시
            st.markdown("---")
            display_pdf_section(selected_date, sel_dept, tab_prefix=f"mismatch_tab_{sel_dept}")
            
    except Exception as e:
        logger.error(f"display_mismatch_content 오류: {e}", exc_info=True)
        st.error(f"데이터 표시 중 오류가 발생했습니다: {e}")


def display_filter_tab():
    st.header("부서별 통계 (불일치 및 누락 항목)")
    if st.session_state.get('force_refresh', False):
        if hasattr(st, 'cache_data'):
            st.cache_data.clear()
        st.session_state.force_refresh = False

    # 1. 사이드바 기간 설정 확인
    if 'work_start_date' not in st.session_state or 'work_end_date' not in st.session_state:
        st.warning("작업 기간을 설정하세요 (사이드바에서).")
        return
    # 사이드바의 현재 날짜 범위 가져오기
    current_start_date = st.session_state.get('date_input_start')
    current_end_date = st.session_state.get('date_input_end')
    
    if current_start_date and current_end_date:
        st.session_state.work_start_date = current_start_date
        st.session_state.work_end_date = current_end_date

    s3_handler = S3Handler()
    
    # 2. 데이터 관리 버튼들
    col1, col2, col3 = st.columns([2, 1, 1])
    with col2:
        if st.button("📊 날짜별 작업 내용 병합", help="각 날짜별 mismatches.json을 통합하여 mismatches_full.json 생성"):
            # 캐시 강제 클리어
            st.cache_data.clear()
            with st.spinner("날짜별 작업 내용을 병합하는 중..."):
                update_result = s3_handler.update_full_mismatches_json()
                if update_result["status"] == "success":
                    st.success(f"✅ 날짜별 작업 내용 병합 완료! 총 {update_result.get('count', 0)}개 항목")
                    # 병합 후 자동으로 새로고침 플래그 설정
                    st.session_state.force_refresh = True
                else:
                    st.error(f"❌ 병합 실패: {update_result.get('message', '알 수 없는 오류')}")
    
    with col3:
        if st.button("🔄 데이터 새로고침", help="최신 통합 데이터를 S3에서 다시 로드합니다"):
            # 선택 상태 보존을 위해 mismatch_data는 삭제하지 않음
            # 대신 캐시만 클리어하여 최신 데이터 로드
            # if 'mismatch_data' in st.session_state:
            #     del st.session_state.mismatch_data  # 주석 처리
            
                                                        
            # 새로고침 플래그 설정 (rerun 제거)
            st.session_state.force_refresh = True
            st.success("데이터 새로고침이 완료되었습니다.")
    
    # S3에서 기존 통합 데이터 로드 (자동 통합 작업 제거)
    with st.spinner("S3에서 통합 데이터를 로드하는 중..."):
        # 강제 새로고침 플래그 처리 (전산누락 저장 후 자동 업데이트)
        if st.session_state.get('force_refresh', False):
            st.session_state.force_refresh = False
            # 캐시 클리어하여 최신 데이터 로드 보장
            st.cache_data.clear()
            st.info("🔄 전산누락 저장으로 인한 자동 데이터 새로고침")
        
        # 기존 통합 파일만 로드 (통합 작업 없음) - 항상 S3에서 최신 데이터 로드
        df_full = s3_handler.load_full_mismatches()
        if df_full is None or df_full.empty:
            st.info("불일치 또는 누락 데이터가 없습니다.\n\n먼저 '날짜별 작업' 탭에서 데이터를 처리하거나 '날짜별 작업 내용 병합' 버튼을 눌러주세요.")
            return
        
        # S3에서 로드한 데이터 정보 표시
        st.info(f"📊 S3에서 로드된 통합 데이터: {len(df_full)}개 항목")
    
    # 사이드바 날짜 범위에 해당하는 완료 처리 로그만 사용하여 필터링
    try:
        completion_logs = st.session_state.get('completion_logs', [])
        if completion_logs:
            # 사이드바 날짜 범위로 완료 로그 필터링 후 적용
            date_range = (st.session_state.work_start_date, st.session_state.work_end_date)
            before_filter = len(df_full)
            mismatch_df = filter_completed_items(df_full, completion_logs, date_range)
            after_filter = len(mismatch_df)
            logger.info(f"부서별 통계 탭 완료 처리 필터링 (기간: {date_range[0]} ~ {date_range[1]}): {before_filter}개 → {after_filter}개")
        else:
            mismatch_df = df_full.copy()
            logger.info("완료 처리 로그가 없어 필터링을 건너뜁니다.")
    except Exception as e:
        logger.warning(f"부서별 통계 탭 완료 처리 필터링 오류: {e}")
        mismatch_df = df_full.copy()
    
    # 세션 상태 덮어쓰기 방지 - 날짜별 작업 탭의 선택 상태를 보호
    # 대신 로컬 변수로만 사용하여 다른 탭에 영향을 주지 않음
    # st.session_state.mismatch_data = mismatch_df.copy()  # 주석 처리
    # 데이터 준비 완료 메시지 간소화
    st.info(f"📊 데이터 준비 완료: {len(mismatch_df)}개 항목")

    # 3. 데이터 준비 및 검증
    if mismatch_df is None or mismatch_df.empty:
        st.info("불일치 또는 누락 데이터가 없습니다.")
        return

    # 4. 데이터 검증 (이미 완료 처리 필터링이 적용된 상태)
    filtered_df = mismatch_df

    # 5. 날짜 컬럼 변환 및 결측치 제거
    filtered_df['날짜_dt'] = pd.to_datetime(filtered_df['날짜'], errors='coerce')
    before_dropna = len(filtered_df)
    filtered_df = filtered_df.dropna(subset=['날짜_dt'])
    after_dropna = len(filtered_df)
    
    if before_dropna != after_dropna:
        st.warning(f"⚠️ 날짜 변환 실패로 {before_dropna - after_dropna}개 항목 제외됨")
    
    # 6. 사이드바 기간으로 필터링
    mask = (
        (filtered_df['날짜_dt'].dt.date >= st.session_state.work_start_date) &
        (filtered_df['날짜_dt'].dt.date <= st.session_state.work_end_date)
    )
    date_filtered_df = filtered_df.loc[mask].copy()
    
    # ★★★ 추가: 부서명 공백 스트립! ★★★
    date_filtered_df['부서명'] = date_filtered_df['부서명'].astype(str).str.strip()
    
    # ===> 여기에 삽입 <===
    print(date_filtered_df[date_filtered_df['부서명'].str.strip() == "11층병동"])
    st.write(date_filtered_df[date_filtered_df['부서명'].str.strip() == "11층병동"])


    # 기간 필터링 결과 간단 표시
    if not date_filtered_df.empty:
        filtered_date_min = date_filtered_df['날짜_dt'].min().strftime('%Y-%m-%d')
        filtered_date_max = date_filtered_df['날짜_dt'].max().strftime('%Y-%m-%d')
        st.info(f"📊 기간 {st.session_state.work_start_date} ~ {st.session_state.work_end_date}: {len(date_filtered_df)}개 항목")

    if date_filtered_df.empty:
        st.warning("선택한 기간에 해당하는 데이터가 없습니다.")
        return
        
    # 7. 부서 필터 (사이드바 기간으로 필터링된 데이터 기준)
    dept_options = ["전체"] + sorted(date_filtered_df['부서명'].dropna().unique())
    selected_dept = st.selectbox("부서 선택", dept_options, key="filter_dept_select")

    if selected_dept == "전체":
        view_df = date_filtered_df
    else:
        view_df = date_filtered_df[date_filtered_df['부서명'].str.strip() == selected_dept]

    # 10. 최종 컬럼 정리 및 데이터 표시
    display_columns = ['날짜', '부서명', '물품코드', '물품명', '청구량', '수령량', '차이', '누락']
    
    # 데이터프레임 처리 (필터 상태가 변경되었을 때만 재처리)
    current_filter_state = selected_dept
    if 'processed_view_df' not in st.session_state or st.session_state.get('last_filter_state') != current_filter_state:
        st.session_state.processed_view_df = view_df.copy()
        for col in display_columns:
            if col not in st.session_state.processed_view_df.columns:
                st.session_state.processed_view_df.loc[:, col] = ""
        
        # 숫자형 컬럼 처리
        for col in ['청구량', '수령량', '차이']:
            st.session_state.processed_view_df.loc[:, col] = pd.to_numeric(
                st.session_state.processed_view_df.loc[:, col], 
                errors='coerce'
            ).fillna(0).astype(int)
        
        # 날짜 포맷 변환
        st.session_state.processed_view_df.loc[:, '날짜'] = pd.to_datetime(
            st.session_state.processed_view_df.loc[:, '날짜'], 
            errors='coerce'
        ).dt.strftime('%Y-%m-%d')
        
        # 누락 컬럼 처리
        st.session_state.processed_view_df.loc[:, '누락'] = st.session_state.processed_view_df.loc[:, '누락'].fillna('').astype(str)
        
        # 컬럼 순서 정리
        st.session_state.processed_view_df = st.session_state.processed_view_df[display_columns]
        
        # 현재 필터 상태 저장
        st.session_state.last_filter_state = current_filter_state

    # 처리된 데이터프레임 표시
    st.dataframe(
        st.session_state.processed_view_df, 
        use_container_width=True,
        column_config={
            "수령량": st.column_config.NumberColumn(format="%d"),
            "PDF수량": st.column_config.NumberColumn(format="%d")
        }
    )

    # 11. 통계 요약
    st.markdown("---")
    st.subheader("통계 요약")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("미완료 항목 수", len(st.session_state.processed_view_df))
        st.metric("표시된 부서 수", st.session_state.processed_view_df.loc[:, '부서명'].nunique())
    with col2:
        st.metric("기간", f"{st.session_state.work_start_date} ~ {st.session_state.work_end_date}")
        # 전산누락 항목 수 계산
        missing_count = st.session_state.processed_view_df.loc[:, '누락'].str.contains('누락', na=False).sum()
        st.metric("전산누락 품목", missing_count)
        
        # 전산누락 데이터 디버깅 정보 (개발용)
        if missing_count > 0:
            missing_dates = st.session_state.processed_view_df[
                st.session_state.processed_view_df.loc[:, '누락'].str.contains('누락', na=False)
            ]['날짜'].unique()
            st.caption(f"전산누락 발견 날짜: {', '.join(sorted(missing_dates))}")
    with col3:
        # 기본 불일치 vs 전산누락 비율
        total_items = len(st.session_state.processed_view_df)
        basic_mismatch = total_items - missing_count
        st.metric("기본 불일치", basic_mismatch)
        if total_items > 0:
            missing_ratio = (missing_count / total_items) * 100
            st.metric("전산누락 비율", f"{missing_ratio:.1f}%")

    # 12. 엑셀 다운로드
    st.markdown("---")
    if st.button("엑셀로 다운로드"):
        # 사이드바 기간 내의 모든 날짜 사용
        available_dates_in_period = sorted(date_filtered_df['날짜_dt'].dt.strftime('%Y-%m-%d').unique())
        excel_data, file_name = download_department_excel(available_dates_in_period)
        if excel_data:
            st.download_button(
                label="엑셀 파일 다운로드",
                data=excel_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("엑셀 파일 생성 중 오류가 발생했습니다.")


def get_department_pages(date_str, dept_name):
    """특정 날짜와 부서의 PDF 페이지 정보를 가져옵니다."""
    try:
        dept_page_tuples = st.session_state.departments_with_pages_by_date.get(date_str, [])
        return [page for dept, page in dept_page_tuples if dept == dept_name]
    except Exception as e:
        logger.error(f"get_department_pages 오류: {e}", exc_info=True)
        return []


def get_department_items(date_str, dept_name):
    """특정 날짜와 부서의 품목 정보를 가져옵니다."""
    try:
        aggregated_ocr_items = st.session_state.get('aggregated_ocr_items_by_date', {}).get(date_str, {})
        return set(map(str, aggregated_ocr_items.get(dept_name, [])))
    except Exception as e:
        logger.error(f"get_department_items 오류: {e}", exc_info=True)
        return set()


def get_mismatch_items(date_str, dept_name):
    """특정 날짜와 부서의 불일치 항목을 가져옵니다."""
    try:
        if 'mismatch_data' in st.session_state and not st.session_state.mismatch_data.empty:
            mismatch_data = st.session_state.mismatch_data[
                (st.session_state.mismatch_data['날짜'] == date_str) &
                (st.session_state.mismatch_data['부서명'] == dept_name)
            ]
            return mismatch_data.copy()
        return pd.DataFrame()
    except Exception as e:
        logger.error(f"get_mismatch_items 오류: {e}", exc_info=True)
        return pd.DataFrame()

# 완료 항목 관리 탭 표시 함수
def display_completed_items_tab():
    """완료 처리된 항목 관리 탭을 표시합니다."""
    try:
        st.header("완료 처리된 항목")
        
        # S3에서 최신 데이터 로드 버튼 추가
        col1, col2 = st.columns([3, 1])
        with col2:
            if st.button("🔄 S3에서 최신 데이터 로드", help="S3에서 완료 처리 로그를 다시 로드합니다"):
                try:
                    s3_handler = S3Handler()
                    completion_logs_result = s3_handler.load_completion_logs()
                    
                    if completion_logs_result["status"] == "success":
                        st.session_state.completion_logs = completion_logs_result["data"]
                        st.success(f"✅ S3에서 {len(st.session_state.completion_logs)}개 완료 로그를 로드했습니다.")
                        st.rerun()
                    else:
                        st.error(f"❌ S3 로드 실패: {completion_logs_result.get('message')}")
                except Exception as e:
                    st.error(f"❌ 로드 중 오류: {e}")
        
        # ⭐️ 세션 상태만 사용 (main()에서 이미 로드됨)
        completion_logs = st.session_state.get('completion_logs', [])
        
        # 세션에 데이터가 없는 경우 안내 메시지
        if not completion_logs and 'completion_logs' not in st.session_state:
            st.info("완료 처리 로그를 로드 중입니다. 잠시 후 다시 시도해주세요.")
            return
                
        if not completion_logs: # 여기서 파일이 없거나, 있어도 내용이 비었거나, 로드 실패 후 빈리스트가 된 모든 경우 처리
            st.info("완료 처리된 항목이 없습니다.")
            return

        # 로드된 데이터 정보 표시
        if completion_logs:
            dates_in_logs = [log.get('날짜', '') for log in completion_logs if log.get('날짜')]
            if dates_in_logs:
                min_date = min(dates_in_logs)
                max_date = max(dates_in_logs)
                st.info(f"📊 로드된 완료 로그: {len(completion_logs)}개 항목 (날짜 범위: {min_date} ~ {max_date})")
            else:
                st.info(f"📊 로드된 완료 로그: {len(completion_logs)}개 항목 (날짜 정보 없음)")

        # DataFrame 생성
        completed_df = pd.DataFrame(completion_logs)

        # 날짜 형식 변환 (정렬/필터용)
        try:
            if '날짜' in completed_df.columns and completed_df['날짜'].dtype == 'object':
                completed_df['날짜_정렬용'] = pd.to_datetime(completed_df['날짜'], format='%Y-%m-%d', errors='coerce')
                mask = completed_df['날짜_정렬용'].isna()
                if mask.any():
                    completed_df.loc[mask, '날짜_정렬용'] = pd.to_datetime(
                        completed_df.loc[mask, '날짜'],
                        format='ISO8601',
                        errors='coerce'
                    )
        except Exception as e:
            logger.error(f"날짜 변환 중 오류 발생: {e}")
            st.error("날짜 형식 변환 중 오류가 발생했습니다.")
            return

        # 날짜 필터링
        if '날짜_정렬용' in completed_df.columns:
            min_date = completed_df['날짜_정렬용'].min()
            max_date = completed_df['날짜_정렬용'].max()
            date_range = st.date_input(
                "날짜 범위 선택",
                value=(min_date, max_date),
                min_value=min_date,
                max_value=max_date
            )
            if len(date_range) == 2:
                start_date, end_date = date_range
                completed_df = completed_df[
                    (completed_df['날짜_정렬용'] >= pd.Timestamp(start_date)) &
                    (completed_df['날짜_정렬용'] <= pd.Timestamp(end_date))
                ]

        # 부서 필터링
        if '부서명' in completed_df.columns:
            dept_options = ['전체'] + sorted(completed_df['부서명'].unique().tolist())
            selected_dept = st.selectbox("부서 선택", dept_options)
            if selected_dept != '전체':
                completed_df = completed_df[completed_df['부서명'] == selected_dept]

        # 정렬
        if '날짜_정렬용' in completed_df.columns:
            completed_df = completed_df.sort_values('날짜_정렬용', ascending=False)

        # 고유키 컬럼 생성 (날짜_부서명_물품코드)
        completed_df['고유키'] = completed_df.apply(
            lambda row: f"{row['날짜']}_{row['부서명']}_{row['물품코드']}", axis=1
        )

        # 체크박스 상태를 위한 세션 변수
        if 'completed_cancel_check' not in st.session_state:
            st.session_state.completed_cancel_check = {k: False for k in completed_df['고유키']}

        # UI: 체크박스와 함께 행 표시
        st.write("**완료 취소할 항목을 체크하세요:**")
        checked_rows = []
        # 표시할 컬럼
        show_cols = ['날짜', '부서명', '물품코드', '물품명', '청구량', '수령량', '차이', '처리시간']
        if '누락' in completed_df.columns:
            show_cols.append('누락')

        # 테이블+체크박스: 행 단위로
        for idx, row in completed_df.iterrows():
            with st.container():
                col1, col2 = st.columns([0.05, 0.95])
                key = row['고유키']
                checked = col1.checkbox(
                    "완료 취소 선택",   # label 필수
                    key=f"completed_cancel_{key}",
                    value=st.session_state.completed_cancel_check.get(key, False),
                    label_visibility="collapsed"
                )
                if checked:
                    checked_rows.append(key)
                # 데이터 표시 (row[show_cols])
                row_data = " | ".join(str(row[c]) for c in show_cols if c in row)
                col2.markdown(row_data)

        # 완료취소 버튼
        if st.button("선택한 항목 완료 취소(되돌리기)", disabled=(not checked_rows)):
            # 체크된 행만 제외하고 새로 저장
            new_df = completed_df[~completed_df['고유키'].isin(checked_rows)]
            new_logs = new_df.drop('고유키', axis=1).to_dict(orient="records")
            
            # S3Handler 생성 (완료 취소 시에만 필요)
            s3_handler = S3Handler()
            save_result = s3_handler.save_completion_log(new_logs)
            st.session_state.completion_logs = new_logs
            # 체크 상태 초기화
            st.session_state.completed_cancel_check = {k: False for k in new_df['고유키']}
            if save_result.get("status") == "success":
                st.success("선택한 항목의 완료 처리가 취소되었습니다.")
            else:
                st.error("완료 취소 저장 중 오류가 발생했습니다.")

        # 통계 정보 표시
        st.markdown("---")
        st.subheader("통계 정보")
        col1, col2 = st.columns(2)
        with col1:
            st.metric("총 완료 항목 수", len(completed_df))
            if '부서명' in completed_df.columns:
                st.metric("처리된 부서 수", completed_df['부서명'].nunique())
        with col2:
            if '날짜_정렬용' in completed_df.columns:
                st.metric("처리 기간", f"{(completed_df['날짜_정렬용'].max() - completed_df['날짜_정렬용'].min()).days + 1}일")
            if '처리시간' in completed_df.columns:
                latest_process = pd.to_datetime(completed_df['처리시간']).max()
                st.metric("최근 처리 시간", latest_process.strftime('%Y-%m-%d %H:%M:%S'))

    except Exception as e:
        logger.error(f"완료 항목 탭 표시 중 오류 발생: {e}", exc_info=True)
        st.error(f"완료 항목을 표시하는 중 오류가 발생했습니다: {e}")


def update_metadata_with_pdf(s3_handler, date_str):
    """PDF 키가 없는 메타데이터를 수정"""
    try:
        # 1. 현재 메타데이터 로드
        meta_result = s3_handler.load_metadata(date_str)
        if meta_result["status"] != "success":
            return False
            
        metadata = meta_result["data"]
        
        # 2. PDF 키가 없으면 추가
        if "pdf_key" not in metadata:
            # PDF 파일 찾기
            pdf_prefix = f"{S3_DIRS['PDF']}{date_str}/"
            response = s3_handler.s3_client.list_objects_v2(
                Bucket=s3_handler.bucket,
                Prefix=pdf_prefix
            )
            
            if 'Contents' in response:
                # 첫 번째 PDF 파일 사용
                pdf_key = response['Contents'][0]['Key']
                metadata["pdf_key"] = pdf_key
                
                # 메타데이터 저장
                s3_handler.save_metadata(date_str, metadata)
                logger.info(f"메타데이터 PDF 키 추가 성공: {date_str}")
                return True
                
        return False
    except Exception as e:
        logger.error(f"메타데이터 수정 실패: {e}")
        return False

def force_reload_excel_data(s3_handler):
    """엑셀 데이터 강제 리로드"""
    try:
        # 1. 누적 엑셀 파일 다운로드
        excel_key = f"{S3_DIRS['EXCEL']}latest/cumulative_excel.xlsx"
        excel_result = s3_handler.download_file(excel_key)
        
        if excel_result["status"] == "success":
            # 2. 엑셀 데이터 로드
            excel_data = pd.read_excel(io.BytesIO(excel_result["data"]))
            
            # 3. 세션에 저장
            st.session_state.excel_data = excel_data
            logger.info(f"엑셀 데이터 강제 리로드 성공: {len(excel_data)} 행")
            return True
            
        return False
    except Exception as e:
        logger.error(f"엑셀 데이터 리로드 실패: {e}")
        return False

def recalculate_mismatches(s3_handler):
    """불일치 데이터를 재계산하고 날짜별로 S3에 저장 (통합 파일 업데이트 포함)"""
    try:
        if 'excel_data' not in st.session_state or st.session_state.excel_data.empty:
            logger.warning("엑셀 데이터가 없어 불일치 데이터를 계산하지 않습니다.")
            return False
            
        # 데이터프레임 복사본 생성
        excel_df = st.session_state.excel_data.copy()
        logger.info(f"불일치 데이터 재계산 시작: 엑셀 데이터 {len(excel_df)}개 행")
        
        # 불일치 데이터 계산
        mismatch_result = data_analyzer.find_mismatches(excel_df)
        
        if mismatch_result["status"] != "success":
            logger.error(f"불일치 데이터 계산 실패: {mismatch_result['message']}")
            return False
            
        mismatch_data = mismatch_result["data"]
        logger.info(f"초기 불일치 데이터 계산 완료: {len(mismatch_data)}개 항목")
        
        # 물품코드 필터링 제거 (process_files에서 이미 제외됨)
        # 완료 처리 로그 필터링만 수행
        completion_logs = st.session_state.get('completion_logs', [])
        if not mismatch_data.empty and completion_logs:
            before_filter = len(mismatch_data)
            mismatch_data = filter_completed_items(mismatch_data, completion_logs)
            after_filter = len(mismatch_data)
            logger.info(f"완료 처리 필터링: {before_filter}개 → {after_filter}개")
        
        st.session_state.mismatch_data = mismatch_data.reset_index(drop=True)
        
        # 날짜별로 S3에 저장
        if not mismatch_data.empty:
            if pd.api.types.is_datetime64_any_dtype(mismatch_data['날짜']):
                unique_dates = mismatch_data['날짜'].dt.strftime('%Y-%m-%d').unique()
            else:
                unique_dates = pd.to_datetime(mismatch_data['날짜'], errors='coerce').dt.strftime('%Y-%m-%d').unique()
            
            for date_str in unique_dates:
                if pd.isna(date_str) or date_str == 'NaT':
                    continue
                    
                if pd.api.types.is_datetime64_any_dtype(mismatch_data['날짜']):
                    date_data = mismatch_data[
                        mismatch_data['날짜'].dt.strftime('%Y-%m-%d') == date_str
                    ].copy()
                else:
                    date_data = mismatch_data[mismatch_data['날짜'] == date_str].copy()
                
                s3_handler.save_mismatch_data(date_str, date_data)
                logger.info(f"날짜 {date_str} 데이터 저장: {len(date_data)}개 항목")
        
        # 전체 통합 파일 업데이트
        update_result = s3_handler.update_full_mismatches_json()
        if update_result["status"] == "success":
            logger.info(f"전체 통합 파일 업데이트 완료: {update_result.get('count', 0)}개 항목")
            return True
        else:
            logger.error(f"전체 통합 파일 업데이트 실패: {update_result['message']}")
            return False
            
    except Exception as e:
        logger.error(f"불일치 데이터 재계산 중 오류 발생: {str(e)}", exc_info=True)
        return False

class ImageCache:
    def __init__(self):
        self._cache = {}
    
    def get_cache_key(self, date_str, dept_name, page_num):
        return f"{date_str}_{dept_name}_{page_num}"
    
    @lru_cache(maxsize=100)
    def get_image(self, cache_key):
        return self._cache.get(cache_key)
    
    def set_image(self, cache_key, image_data):
        self._cache[cache_key] = image_data

def process_images_parallel(images: List[Dict], max_workers: int = 4):
    """이미지 처리를 병렬로 수행"""
    results = []
    
    def process_single_image(img_info):
        try:
            img_bytes = get_pdf_preview_image_from_s3(img_info["file_key"])
            return {
                "status": "success",
                "data": img_bytes,
                "info": img_info
            }
        except Exception as e:
            logger.error(f"이미지 처리 중 오류: {e}")
            return {
                "status": "error",
                "message": str(e),
                "info": img_info
            }
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_image = {
            executor.submit(process_single_image, img_info): img_info 
            for img_info in images
        }
        
        for future in concurrent.futures.as_completed(future_to_image):
            result = future.result()
            results.append(result)
    
    return results

@st.cache_data(ttl=3600) # 1시간 캐시
def get_preview_images_for_s3(date_str, s3_handler_dirs, s3_handler_bucket, s3_handler_aws_config):
    s3_config_temp = {
        "aws_access_key_id": s3_handler_aws_config["aws_access_key_id"],
        "aws_secret_access_key": s3_handler_aws_config["aws_secret_access_key"],
        "region_name": s3_handler_aws_config["region_name"]
    }
    s3_client_temp = boto3.client('s3', **s3_config_temp)
    
    # 임시 S3Handler 유사 객체 또는 직접 s3_client_temp 사용
    # 여기서는 s3_handler.load_metadata 호출을 모방
    metadata_key = f"{s3_handler_dirs['METADATA']}{date_str}/metadata.json"
    try:
        response = s3_client_temp.get_object(Bucket=s3_handler_bucket, Key=metadata_key)
        metadata = json.loads(response['Body'].read())
        return metadata.get("preview_images", [])
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            return [] # 메타데이터 없으면 빈 리스트
        logger.error(f"메타데이터 로드 실패 ({date_str}) in get_preview_images_for_s3: {e}")
        return [] # 오류 시 빈 리스트
    except Exception as e:
        logger.error(f"get_preview_images_for_s3 예외 ({date_str}): {e}")
        return []


def get_all_dept_images_for_dates(dates_to_load_tuple, selected_dept_filter, s3_handler_dirs, s3_handler_bucket, s3_handler_aws_config):
    all_dept_images = {}
    for date_str in dates_to_load_tuple:
        # 캐싱된 함수를 호출하여 특정 날짜의 preview_images를 가져옴
        preview_images = get_preview_images_for_s3(date_str, s3_handler_dirs, s3_handler_bucket, s3_handler_aws_config)
        for img_info in preview_images:
            dept = img_info.get("dept")
            if not dept: continue
            if selected_dept_filter == "전체" or dept == selected_dept_filter:
                if dept not in all_dept_images:
                    all_dept_images[dept] = []
                img_info['date'] = date_str 
                all_dept_images[dept].append(img_info)
    return all_dept_images

if __name__ == "__main__":
    # S3 연결 확인
    if not check_s3_connection():
        st.error("S3 스토리지 연결에 실패했습니다. 관리자에게 문의하세요.")
        exit()
    
  
    main() 